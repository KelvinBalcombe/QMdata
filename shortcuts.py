# -*- coding: utf-8 -*-
"""
Created on Sat May 20 08:48:19 2017

@author: aes05kgb
"""
def pbold(term):
    x=''
    x='\033[1;31m' +  term + '\033[0m'
    return  x

def helpme():
    print('In order to find help on a topic you need to input a number. The help output in red can be copied and pasted into a cell and run without alteration as an example')
    n=input('FOR HELP ON: \n Changing the size of plots, enter 1  '
            +'\n Manipulating Fractions, enter 2'
            +'\n Simultaneous equations, enter 3'
            +'\n Plotting a line through 2 points, enter 4 '
            +'\n Plotting and solving demand and supply, enter 5'
            + '\n Plotting and solving breakeven problems, enter 6'
            +'\n Plotting polynomial equations, enter 7'
            +'\n Discrete Probability Distributions, enter 8'
            +'\n Continuous Probability Distributions enter 9'
            +'\n General Statistics Help using the menu command, enter 10 '
            +'\n Discounting and Present Value, enter 11'
            +'\n Defining Functions, enter 12' 
            +'\n Differentiation, integration or optimisation, enter 13' 
            +'\n Matrices, enter 14' 
            +'\n'  )
    if n==str(1): 
        print('\nTo make plots larger or smaller use \n'+ 'pltsize(verticalsize,horizontalsize) \n e.g. to give vertical depth of 8 and horzontal length of 12 use:\n' + pbold('pltsize(8,12)'))
    elif n==str(2):
        print('\nTo define a fraction of a over b use frac(a,b). e.g. 1/2 is ')
        print(pbold('frac(1,2)'))
        print('If you add substract or multiply fractions they will remain fractions in this format e.g.')
        print(pbold('frac(1,2)+frac(2,3)') +' or '+pbold('frac(1,2)*frac(2,3)')  )      
    elif n==str(3):
        print('\nTo plot and solve simultaneous equations with the slopes and intercepts use: ')
        print('plotsimult(intercept1,slope1,intercept2,slope2). e.g.  ')
        print(pbold('   plotsimult(1,2,3,-1)'))
    elif n==str(4):
        print('\nTo plot a line through two points [x0,y0]and [x1,y1] use plotlinefrompoints([x0,y0],[x1,y1]).  e.g.')
        print(pbold('plotlinefrompoints([1,1],[2,2])'))
    elif n==str(5):
        print('\n To plot and solve supply and demand equations with specified slopes and intercepts use:')
        print('   plotSD(supplyint,supplyslope,demandint,demandslope,loglinear=False)')
        print(' -note that the demand slope must be positive and the loglinear can be changed to "True" for log linear equations. e.g.')
        print(pbold('plotSD(1,1,10,-1)'))
    elif n==str(6):
        print('\nTo solve breakeven problems with costs and revenues use: \n'+ 'plotbreakeven(fixedcost,marginalcost,marginalrevenue)' +'\n  - note that marginal revenues is equal to the price. e.g. \n')
        print(pbold('plotbreakeven(100,2,3)'))
    elif n==str(7):
        print('\n To plot polynomial equations and their slopes use e.g.')
        print(pbold('   plotfunction(0,c1=1,p1=1,c2=1,p2=2,c3=1,p3=1)')+'\n where you change the coefficients c and powers p\n' )        
    elif n==str(8):   
        print(' \nTo obtain the probability distribution for a given number of outcomes choose distmap(distribution,value)')
        print(' The distribution need to be specified as'+' binom(n,p), nbinom(n,p) or poisson(n)')
        print(' e.g. the probabilities associated with 4 successes in 10 trials with probability of success in each trial being 0.5 is \n'+ pbold('distmap(binom(10,0.5),4)'))
        print(' e.g. the probabilities associated with 4 failures to get 10 successes (thus 14 trials) where the probability of success in each trial is 0.5 is \n'+ pbold('distmap(nbinom(10,0.5),4)'))
        print(' e.g. the probabilities associated with 4 events in a Possion experiment where the mean number of events is 4 \n'+ pbold('distmap(poisson(10),4)'))
        
    elif n==str(9):
        print('\nTo obtain the probability of getting values more then or less than a given value use distcap(norm(mean,stdv),lowerval,upperval)')
        print('e.g. form a normal distribution with mean 2 and standard deviation of 1 the probabilities of getting values of less than 0 and more than 1.6 are obtained using \n' + pbold('distcmap(norm(2,1),0,1.6)') )
    elif n==str(10):
        '\n' 
        statshelp()
    elif n==str(11):
        print('\nBelow y represents a series of payments and revenues and r represents the interest rate of 5%\n'+pbold('presentval(y=[-500,200,150,100,50,40,30],r=0.05,plot=False)') +'\n Changing plot to True will give a plot')
    elif n==str(12):
        print('\n To specify a function you need a statement such as def f(x): return somefunctionofx  for example if we want to define the function f(x) that is the square of x we would specify')
        print(pbold('def f(x): return x**2' )) 
        print('If you then input a number it will return the value of the function at that point e.g')
        print(pbold('f(2)'))
        print('Note that if you replace f() with g() or h() then this is perfectly fine') 
        print('Functions can be for more than one argument e.g. the following will square x and and the square of z')
        print(pbold('def f(x,z): return x**2+z**2' ))   
        print('If you then input a numbers it will return the value of the function at that point e.g')
        print(pbold('f(2,3)'))      
    elif n==str(13) or n==str(14):
        print('\nAt the moment you have imported the shortcuts commands. To get python to do differentiation, integration, optimisation or matrices you need to import the turnonsympy file. To do this you must initiate a new python file first and run ' + pbold('from turnonsympy import *') +  ' as the first command in the first line. Once you have done this you can use the '+ pbold('helpme()')+ ' command in the new file to get help with plotting and differention, optimisation or matrices using the sympy options') 
            
        
    return

def statshelp():
    print( '\033[1m'+'To load in a dataframe called "data" you need to type:'
          +'\033[0m' + '\033[1;31m'+'\ndata=load("full name of file")'+'\033[0m' 
          + '\nThe name of the file needs to include the directory and the suffix of the filename if it is not in your jupyter QM directory. \nHowever, if the file is loaded in your jupyter QM directory it will not need a directory name'+
'\033[1m'
        +' \nThe following command initiate menus that then employ the options or tests listed under "direct commands" in the "Quick List of Commands" File , where it is assumed that "data" is a valid data frame i.e. a dataframe named data\n'+'\033[0m'
        +'\033[1;31m'+'variableview(data)'+'\033[0m'+ ' will list the variables in the data set along with their type'
        +'\033[1;31m'+'\nt_tests(data)'+'\033[0m'+' will start a menu do to t-tests about means'  
        +'\033[1;31m'+'\nnp_tests(data)'+'\033[0m'+' will start a menu to do non-parametric tests about medians' 
        + '\033[1;31m'+'\np_tests(data)'+'\033[0m'+'  will start a menu to do tests of proportions'    
        +'\033[1;31m'+ '\nu_tests(data) '+'\033[0m'+'  will start a menu to do uniformity tests'    
        +'\033[1;31m'+ '\nv_tests(data)  '+'\033[0m'+' will start a menu to test for variances'    
        +'\033[1;31m'+ '\ncontables(data)'+'\033[0m'+' will start a menu to do contingence tables'    
        +'\033[1;31m'+ '\nnorm_tests(data)'+'\033[0m'+' will start a menu to do test for normality'    
        +'\033[1;31m'+ '\nmean_CIs(data)'+'\033[0m'+' will start a menu to construct confidence intervals about the mean'    
        +'\033[1;31m'+ '\ncorrs(data) '+'\033[0m'+'will start a menu to look at bivariate correlations'    
        + '\033[1;31m'+'\nregression(data)'+'\033[0m'+' will start a menu to conduct regressions and Anovas'    
        +'\033[1;31m'+ '\nprobit(data) '+'\033[0m'+'will start a menu to conduct binomial probit regressions'    
        + '\033[1;31m'+'\nordprobit(data)'+'\033[0m'+' will start a menu to conduct ordinal probit regressions'  
        + '\033[1m'+  '\nA number of the options above have plot options however, there are some pure interactive plot options as below '+'\033[0m'
        + '\033[1;31m'+'\nscatterplot(data)'+'\033[0m'+' will start a menu to do scatter plots'    
        + '\033[1;31m'+'\nboxplot(data) '+'\033[0m'+'    will start a menu to do box plots'    
        + '\033[1;31m'+'\nhistogram(data)'+'\033[0m'+'   will start a menu to do histograms'    
        + '\033[1;31m'+'\ncounts(data) '+'\033[0m'+'     will start a menu plot counts of data  (or proportions)' 
        + '\033[1m   '+' \nThe following interactive command for pivot-tables is also included' 
        + '\033[1;31m'+ '\npivot(data)'+'\033[0m'
        + '\033[1m   '+'\nThe following are non-interactive but simple and useful:' +'\033[0m' 
        + '\033[1;31m'+ '\ndescribe(data)'+'\033[0m'+'  will describe the data in "data"' 
        + '\033[1;31m'+ '\ncronbachalpha(data) '+'\033[0m'+' will compute the cronbachalpha for all variables in "data"') 
    return
      

from sympy import Rational as frac

from ipywidgets import interact, interactive, fixed, interact_manual


#If you want to set up a spreadsheet that you pull and push data too by default give it the name here
excelname='C:\\Users\\aes05kgb\\Desktop\\pythondata.xlsx'
#datadirec="https://kelvin1.z13.web.core.windows.net//"
datadirec="https://raw.githubusercontent.com/KelvinBalcombe/python/master/"
datadirec2='c:\\k\\qm resources\\Data sets\\'
plotdefault=True

Alphabet=['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']

import warnings
warnings.filterwarnings("ignore")

import os
import glob
def myfiles(ext=''):
    if ext=='':
        return os.listdir()
    else:
        return glob.glob( '**.' +str(ext) )

from webbrowser import open as openweb

import sys
from sys import exit as stop

import numpy as np

from numpy import arange,linspace,isnan
from numpy.linalg import inv,cholesky,det,slogdet
from numpy import concatenate,shape,kron,exp,array,pi,round,floor,ceil,square,sqrt,power,copy,empty,zeros_like,squeeze,matrix
from numpy import log as ln
from numpy import log10 as log
from numpy import multiply as mult_
from numpy import divide as div_
from numpy.random import multivariate_normal as rndmn_
from numpy.random import normal as rndn_
from numpy.random import uniform as rndu_
from numpy.random import beta as rndb_
from numpy.random import standard_t as rndt_
from numpy.random import gamma as rndg_
from numpy.random import randint as rndint_
from numpy.random import multinomial as rndmult_

scipyshort=True
if scipyshort:    
    import scipy as sp
    import scipy.stats as sps
    
    from scipy.special import gamma as gammaf
    from scipy.special import gammaln as lngammaf
    #Distributions
    from scipy.stats import norm,beta,binom,poisson,nbinom,gamma,loggamma,lognorm,uniform,rankdata
    from scipy.stats import multivariate_normal as mvn
    from scipy.stats import t as tdist
    from scipy.stats import chi2 as chidist
    from scipy.stats import f as fdist
    from scipy.optimize import minimize
    
    #Stat tests
    from scipy.stats import ttest_ind,ttest_1samp,chi2_contingency    
    from scipy.stats import describe as describe_
    #from scipy.optimize import curve_fit as nls
    from scipy.cluster.vq import kmeans2
    #from scipy.cluster.hierarchy import dendrogram, linkage


import pandas as pd
from pandas import DataFrame as frame, pivot_table
from pandas import get_dummies
from pandas import read_excel
from pandas import date_range
from pandas import datetime
#import pandas_datareader.data as web


import matplotlib.pyplot as plt
import matplotlib.pylab as pylab
params = {'axes.titlesize':'x-large',
          'legend.fontsize':'x-large'}
pylab.rcParams.update(params)

def pltsize(x1=10,x2=16):
    plt.rcParams['figure.figsize'] = [x2, x1]
    return

import pickle

import patsy
from patsy import dmatrix,dmatrices,ModelDesc

from statsmodels.api import OLS,add_constant
import statsmodels.stats.api as sms
from statsmodels.stats.api import anova_lm as anova
import statsmodels.formula.api as smf
from statsmodels.stats.proportion import proportions_ztest as proptest
from statsmodels.formula.api import ols,probit,glsar,quantreg,mixedlm,rlm
from statsmodels.base.model import GenericLikelihoodModel
from statsmodels.stats.diagnostic import het_breuschpagan as breushpagan
from statsmodels.stats.diagnostic import acorr_breusch_godfrey as breushgodfrey


def integers(start=1,stop=26):
    return arange(start,stop+1)

def alphabet(start=1,stop=26):
    return array(Alphabet[start-1:stop])

def cwd(level=0):
    if level==0:
        return os.getcwd()
    elif level==-1:
        return os.path.split(os.getcwd())[0]
    
def sopen():
    return os.startfile(excelname)

def pull(s='Sheet1'):
    if type(s) ==int:
       s='Sheet'+str(s)
    return load(excelname,sheetname=s)

def push(tab,s='output'):
    tab=frame(tab)
    if type(s) ==int:
       s='Output'+str(s)
    return addtoexcel(tab,excelname,sheetname=s,startrow=0,startcol=0,rounds=0)

#Removes white spaces from columns of a data set
def nospace(data):
    datas=data.copy()
    cols=[]
    for i in list(datas):
        cols=cols+[i.replace(' ','')]
    datas.columns=cols
    return datas

def lowercase(data):
    datas=data.copy()
    cols=[]
    for i in list(datas):
        cols=cols+[i.lower()]
    datas.columns=cols
    return datas

def shorten(data,d=1,mapnames=True):
    datas=data.copy()
    z=list(datas)
    #print(z)
    k=[]
    for i in z:
            #print(i.split(' '))
            k=k+[len(i.split(' '))]
    #print(k)
    cols=[]
    j=0
    for i in z:
        if d==1 or k[j]==1:
            #print('yes')
            cols=cols+[i.split(' ')[0]]
        elif d==2 and k[j]>1:
            #print(i,i.split(' '))
            cols=cols+[i.split(' ')[0]+i.split(' ')[1]]
        j=j+1
    
    ii=0
    for i in cols:
        jj=0
        for j in cols:
            ad=2
            if i==j and jj !=ii:
               #print('repeat',i,j)
               j=j+str(ad)
               ad=ad+1 
               cols[jj]=j 
            jj=jj+1
            
        ii=ii+1
            
    datas.columns=cols
    oldnew=mapping(data,datas)
    if mapnames==False:
        return datas
    else:
        print(oldnew)
        return datas

def mapping(data1,data2):
    d1=list(data1.columns)
    d2=list(data2.columns)
    map=frame([d1,d2]).T
    map.columns=['old','new']
    map.index=map['old']
    del map['old']
    return map

def load(name,sheetname=0,header=0,nospace=False,lowercase=False,printhead=True):
    if str('http') in name:
        name=name.replace(" ","%20")
    s=name.split('.')
    print(s)
    if len(s)>1:
        if name.split('.')[len(s)-1]=='npy':
            z=np.load(name)
        elif name.split('.')[len(s)-1]=='dta':
            z=pd.read_stata(name)
        elif name.split('.')[len(s)-1]=='xls':
            z=pd.read_excel(name,sheet_name=sheetname,header=header) 
            if printhead:
                print(z.head())
        elif name.split('.')[len(s)-1]=='xlsx':
            z=pd.read_excel(name,sheet_name=sheetname,header=header) 
            if printhead:
                print(z.head())
        elif name.split('.')[len(s)-1]=='csv':
            z=pd.read_csv(name)    
        if nospace:
           z=nospace(z) 
        if lowercase:
           z=lowercase(z)  
        
    else:
        z=pd.read_pickle(name)
    return z

def vec(A):
   return np.reshape(np.transpose(A), (-1,1))

def save(z,name='temp'): #If numpy object, you do not need the 'npy'
    if type(z)==pd.core.frame.DataFrame:
        z.to_pickle(name)
    else:
        np.save(name,z)
    return 

def addtoexcel(tab,name,sheetname,startrow=0,startcol=0,rounds=0):
    from openpyxl import load_workbook
    #wb = openpyxl.Workbook()
    
    book = load_workbook(name)
    writer = pd.ExcelWriter(name, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    if rounds !=0:
        tab.round(rounds).to_excel(writer, sheetname, startrow=startrow, startcol=startcol)
    else:
        tab.to_excel(writer, sheetname, startrow=startrow, startcol=startcol)
    try:    
        writer.save()
    except:
        writer.write()
        pass
    return

def toexcel(tab,name,sheetname='Sheet',startrow=0,startcol=0,rounds=0,new=True):
    from openpyxl import load_workbook,Workbook
    if new:
        book= Workbook()
        book.save(name)
    
    book = load_workbook(name)
    writer = pd.ExcelWriter(name, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    if rounds !=0:
        tab.round(rounds).to_excel(writer, sheetname, startrow=startrow, startcol=startcol)
    else:
        tab.to_excel(writer, sheetname, startrow=startrow, startcol=startcol)
    try:    
        writer.save()
    except:
        writer.write()
        pass
    return


def rndint(ob3,ob1=0,ob2=1):
    #ob1 is the low and ob2 is the high
    #ob3=checkob(ob3)
    return (twodma(rndint_(ob1,ob2+1,ob3)))

def rndn(ob3,ob1=0,ob2=1):
    ob3=checkob(ob3)
    return (rndn_(ob1,ob2,ob3))

def rndu(ob3,ob1=0,ob2=1):
    ob3=checkob(ob3)
    return rndu_(ob1,ob2,ob3)

def twodm(x): 
    s=shape(x)
    if len(s)>2:
        x=np.squeeze(x)
        s=shape(x)    
    if len(s)==1:
        x=x.reshape(s[0],1)
    if len(s)==0:
        x=x.reshape(1,1)
    return x

#This will give 2.d shape to the array, preserving 2.d. shape if it exists, otherwise a column vector
def twodma(x):
    if type(x) !=np.ndarray:
        x=np.array(x)
    x=twodm(x)
    return x
   
#This will give 2.d. to the the array, but enforcing a column array, or row array if one of the dimensions is 1
def twodmca(x,column=True):
    x=squeeze(twodma(x))
    s=shape(x)
    if len(s)==0:
        x=x.reshape(1,1)
    if len(s)==1:
        if column:
            x=x.reshape(s[0],1)
        else:
            x=x.reshape(1,s[0])
    return x    
    
def matrix(x): #If it is not already a matrix, change to a matrix, always a column vector if one dimensional
    if type(x) != np.matrixlib.defmatrix.matrix:
        x=twodma(x)
        x=np.matrix(x)
    return x

def checkob(ob):
    if type(ob)==list:
        if len(ob)==1:
           ob=[ob[0],1] 
    if type(ob)==int:
        ob=[ob,1]
    return ob
    
def ones(ob):
    ob=checkob(ob)
    return twodm(np.ones(ob))

def zeros(ob):
    ob=checkob(ob)
    return twodm(np.zeros(ob))
    
#def corr(x):
    return twodm(np.corrcoef(x.T))    

def cov(x):
    return twodm(np.cov(x.T))    

def rows(x):
    return shape(twodma(x))[0]

def cols(x):
    return shape(twodma(x))[1]
    
def sumc(x):
    return twodm(np.sum(x,axis=0))

def cumsumc(x):
    return twodm(np.cumsum(x,axis=0))

def pivot(data,groupby='',y='',summary='mean'):
    return pivot_table(data,index=groupby,values=y,aggfunc=summary)

def vecdf(df,indexfrom1=True):
    vdf=vec(twodma(df))
    group=list()
    j=0
    for i in range(1,rows(vdf)+1):
        group=group+[df.columns[j]]
        if i%rows(df)==0 and i !=1:
            j=j+1      
    vdf=frame(vdf) 
    group=frame(group)
    newdf=cc([vdf,group])
    newdf.columns=['values','group']
    if indexfrom1:
        newdf=from1(newdf)
    return newdf

def unvecdf(vdf,group='group',values='values',indexfrom1=True):
    g=set(vdf[group])
    rws=int(rows(vdf)/len(g))
   
    v=twodma(vdf[values])
    
    df=frame(reshape(v,rws,len(g)))
    lst=list()
    for i in vdf[group]:
        if i not in lst:
            lst=lst+[i]
    df.columns=lst
    if indexfrom1:
        df=from1(df)
    return df

def dropthenas(Y,X=0):
    try:
        Y=frame(Y)
        Y=Y.dropna()
    except:
        pass
    try:
        X=frame(X)
        X=X.dropna()
    except:
        pass
    return Y,X

def makenormal(x):
    x=frame(x)
    m=meanc(twodma(x))
    s=stdc(twodma(x))
    y=x.rank()/(rows(x)+1)
    z=-norm.isf(y)
    z=m+s*twodma(z)
    #z=frame(z)
    #z.index=y.index
    #z.columns=y.columns
    return twodma(z)

def svar(X):
    n = float(len(X))
    svar=(sum([(x-np.mean(X))**2 for x in X]) / n)* n/(n-1.)
    return svar

def kclust(data,y='',k=2):
    if y=='':
       y=data.columns
    dat=data[y].dropna()
    out=kmeans2(dat.astype(float),k)
    out1=frame(out[1],index=dat.index,columns=['cluster'])
    out0=findex(frame(out[0],columns=[dat.columns]))
    Z=cc([data,out1])
    return [out0,Z]

def cronbachalpha(data,y=''):
    if y=='':
       y=data.columns
    datas=data[y]
    itemscores=twodma(datas).T
    itemvars = [svar(item) for item in itemscores]
    #print(itemvars)
    tscores = [0] * len(itemscores[0])
    for item in itemscores:
       for i in range(len(item)):
          tscores[i]+= item[i]
    nitems = len(itemscores)
    #print "total scores=", tscores, 'number of items=', nitems

    Calpha=nitems/(nitems-1.) * (1-sum(itemvars)/ svar(tscores))
    print("Cronbach alpha = ", Calpha)
    Calpha=float(Calpha)
    if Calpha >=0.9:
            print('Excellent')
    elif Calpha<0.9 and Calpha>=0.8:
            print('Good')    
    elif Calpha<0.8 and Calpha>=0.7:
            print('Acceptable')
    elif Calpha<0.7 and Calpha>=0.6:
            print('Questionable')
    elif Calpha<0.6 and Calpha>=0.5:
            print('Poor')                 
    else:
            print('Unacceptable')
    return Calpha


def m_w_test_(y,x=0,plot=plotdefault): 
    diff=False
    if rows(x)>1:
        diff=True
        
    out=sps.wilcoxon(squeeze(y-x))
    t=out[0]
    p=out[1]
    
    if diff:
        x=0
    z=frame([np.median(y),x,float(t),float(p)], index=['Median','H0:Symmetric around','w-stat','P-val'],columns=['Rank Sign Test'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        title= 'HO: Distribution is symmetric around ' + str(x)
        plotbarpval(p,title=title,ax=ax[0]) 
        plotpiepval(p,title='',ax=ax[1])
    return z.round(6).T

def m_w_test(data,y='y',x=0,plot=plotdefault):
    y=data[y]
    if type(x)==str:
        x=data[x]
    z=m_w_test_(y=y,x=x,plot=plot)
    return z



def v_chi_test_(Y,X=1,plot=plotdefault):
    X=sqrt(X)
    s1=stdc(twodma(Y))
    n1=rows(Y)
    stat=(n1-1)*(s1/X)**2
    pval1=chidist.cdf(stat,n1-1)
    pval2=(1-chidist.cdf(stat,n1-1))
    t=stat
    p=2*np.min([pval1,pval2])           
    z=frame([X**2,float(t),float(p)], index=['H0: Var=','chi-stat','P-val'],columns=['Chi-Square Test'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        title= 'HO: Variance is equal to ' + str(X**2)
        plotbarpval(p,title=title,ax=ax[0]) 
        plotpiepval(p,title='',ax=ax[1])
    return z.round(6).T

def v_chi_test(data,y='y',x=1,plot=plotdefault):
    y=twodma(data[y])
    #print(y)
    return v_chi_test_(y,X=x,plot=plot)

def u_chi_test_(y,plot=plotdefault):
    x=twodma(y)
    z=twodma(x-meanc(x))
    y_=sumc(z**2)/meanc(x)
    #print(len(x))
    out=[float(y_),float(1-chidist(len(x)-1).cdf(y_))]
    t=out[0]
    p=out[1]
    z=frame([t,p], index=['chisquare','P-val'],columns=['H0: Uniform Distribution'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        title= 'HO:  Uniform Distribution'
        plotbarpval(p,title=title,ax=ax[0]) 
        plotpiepval(p,title='',ax=ax[1])
        plt.show()
    return z.T.round(6)

def u_tests(data):
    u=['']+list(data.columns)
    return interact_manual(u_chi_test,df=fixed(data),y=u)

def norm_ks_test_(y,plot=plotdefault):
    y=zscore(twodma(y))
    out=sps.kstest(squeeze(y),'norm')
    t=float(out[0])
    p=float(out[1])
    z=frame([t,p], index=['KS-Stat','P-val'],columns=['Normality Test'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        title= 'HO:  Normal Distribution'
        plotbarpval(p,title=title,ax=ax[0]) 
        plotpiepval(p,title='',ax=ax[1])
        plt.show()
    return z.T.round(6)

def norm_tests(data):
    u=['']+list(data.columns)
    return interact_manual(norm_ks_test,df=fixed(data),y=u,groupby=u)

def norm_ks_test(df,y='y',groupby='',group='',plot=plotdefault):
    if groupby !='':
        y=df[df[groupby]==group][y].astype(float)
    else:    
        y=df[y].astype(float)
    y=y.dropna()
    return norm_ks_test_(y,plot)

def stackdf(df):
    return vecdf(df).dropna()

def unstackdf(vdf,y='y',groupby='group',indexfrom1=True):
    g=set(vdf[groupby])
    j=1
    for i in g:
        s=vdf[groupby]==i
        dfi=vdf[s]
        dfi=dfi[y]
        dfi=frame(dfi)
        dfi.columns=[i]
        dfi=findex(dfi)
        if j==1:
           df=dfi.copy()
           #print(df)
           #stop()
        else:
           df=pd.concat([df,dfi],axis=1) 
        j=j+1
    return df

def v_l_test_(y,plot=plotdefault):
    y=twodma(y).astype(float)
    if cols(y)==2:
        out=sps.levene(y[:,0],y[:,1])
    elif cols(y)==3:   
        out=sps.levene(y[:,0],y[:,1],y[:,2])
    elif cols(y)==4:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3]) 
    elif cols(y)==5:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3],y[:,4])     
    elif cols(y)==6:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3],y[:,4],y[:,5])
    elif cols(y)==7:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3],y[:,4],y[:,5],y[:,6])    
    elif cols(y)==8:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3],y[:,4],y[:,5],y[:,6],y[:,7])   
    elif cols(y)==9:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3],y[:,4],y[:,5],y[:,6],y[:,7],y[:,8])   
    elif cols(y)==10:   
        out=sps.levene(y[:,0],y[:,1],y[:,2],y[:,3],y[:,4],y[:,5],y[:,6],y[:,7],y[:,8],y[:,9])       
    t=out[0]
    p=out[1]
    z=frame([t,p], index=['Levene-stat','P-val'],columns=['H0:Common Variance'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        title= 'HO:  Equal Variances'
        plotbarpval(p,title=title,ax=ax[0]) 
        plotpiepval(p,title='',ax=ax[1])
    return z.T

def v_l_test(df,y='y',groupby='',plot=plotdefault):
    if groupby !='':
        y=unstackdf(df,y=y,groupby=groupby)
    else:    
        y=df[y]
    y=y.dropna().astype(float)
    return v_l_test_(y,plot)

def u_chi_test(df,y='y',plot=plotdefault):
    df=df[y]
    df=df.dropna()
    df_col1,names1 = stringtonumber(df)
    #print(names1)   
    result = [[sum((df_col1 == cat1)) for cat1 in categories(df_col1)]]
    result=frame(result)
    result.columns=names1
    result.index=[y]
    #print(result)
    out=u_chi_test_(twodma(result.T),plot)  
    return out.round(6)


def p_bnm_test_(y=1,n=1,p=0.5,plot=plotdefault,proportions=False):
    x=y
    if x<1:
       proportions=True
    if proportions:
       x=round(x*n)
    pval=sps.binom_test(x=x, n=n, p=p)
    out=frame([p,x,x/n,n,pval],index=['H0:p=','Counts','Proportion','Trials','P-val'],columns=['Binom-Test'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        title= 'HO: probability is equal to ' + str(p) 
        plotbarpval(pval,title=title,ax=ax[0]) 
        plotpiepval(pval,title='',ax=ax[1])
        plt.show()
    return out.T

def p_bnm_test(data,y='y',groupy=1,p=0.5,plot=plotdefault):
    n=rows(data)
    x=float(sumc(data[y]))
    return p_bnm_test_(y=x,n=n,p=p)

def p_z_test_(y,n,p=0,plot=plotdefault,proportions=False):
    x=twodma(y)
    if x[0]<1:
       proportions=True
    n=twodma(n)
    if proportions:
       x[0]=n[0]*x[0]
       if rows(x)>1:
           x[1]=n[1]*x[1]
    if type(x)==float or type(x)==int:
        x=[x]
    t,pval=proptest(x,n,p,prop_var=p) ;
    t=t[0]
    pval=pval[0]
    if len(x)==1:
        z=frame([float(p),x,(x/n),n,float(t),float(pval)], index=['H0:p= ','Counts','Proportions','N','Zval','P-val'],columns=['Z-Test'])
    else:
        z=frame([float(p),x,(x/n),n,float(t),float(pval)], index=['H0:diff in p= ','Counts','Proportions','N','Zval','P-val'],columns=['Z-Test'])
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        if len(x)==2:
           title= 'HO: proportions are the same'
        else:
           title= 'HO: proportion is equal to ' + str(p) 
        plotbarpval(pval,title=title,ax=ax[0]) 
        plotpiepval(pval,title='',ax=ax[1])
        
        plotdist(t,tail='s',distr=norm,mean=0,stdv=1,value=True)
        plt.show()
    return z.round(6).T

def p_z_test(df,y='y',x='',mu='',groupy='',groupx='',plot=False,verbose=False):
    if verbose:
        if groupy =='':
                print('You need to specify which group within y you wish to test the proportion by setting groupy')
            
        if mu !='':
            print('mu is active, therefore all other settings are ignored (i.e. as if they were blank). This test will test whether the portion of groupy in y is mu')
            groupx=''
            x=''
            
        if mu=='' and x=='':
                print('You need either to specify mu or x') 
        
        if x !='' and groupx =='':
                print('You need to specify which group within x you wish to test the proportion by setting groupx')
        
    p1=proportion(df,y,groupy)
    if x != '' and type(x) !=float and type(x) != int:
        p2=proportion(df,x,groupx)
        return p_z_test_([p1,p2],[rows(df),rows(df)],plot=plot)
    else:
        mu=float(mu)
        return p_z_test_(p1,rows(df),mu,plot=plot)
    
def p_z_testi(df,y='y',x='',mu='',groupy='',groupx='',plot=False,verbose=False):
    if verbose:
        if groupy =='':
                print('You need to specify which group within y you wish to test the proportion by setting groupy')
            
        if mu !='':
            print('mu is active, therefore all other settings are ignored (i.e. as if they were blank). This test will test whether the portion of groupy in y is mu')
            groupx=''
            x=''
            
        if mu=='' and x=='':
                print('You need either to specify mu or x') 
        
        if x !='' and groupx =='':
                print('You need to specify which group within x you wish to test the proportion by setting groupx')
        
    p1=proportioni(df,y,groupy)
    if x != '' and type(x) !=float and type(x) != int:
        p2=proportioni(df,x,groupx)
        return p_z_test_([p1,p2],[rows(df),rows(df)],plot=plot)
    else:
        mu=float(mu)
        return p_z_test_(p1,rows(df),mu,plot=plot)
    
    
def p_tests(data,y='y',mu='',x='',groupy='',groupx=''):
    u=['']+list(data.columns)
    return interact_manual(p_z_testi,df=fixed(data),verbose=fixed(True),y=u,x=u)    

def m_mww_test_(y,x,plot=plotdefault):
    res=sps.mannwhitneyu(y, x, use_continuity=True, alternative=None)
    if plot==True:
        fig,ax=plt.subplots(ncols=2)
        if len(x)==2:
           title= 'HO: medians are the same'
        else:
           title= 'HO: medians are not the same' 
        plotbarpval(res[1],title=title,ax=ax[0]) 
        plotpiepval(res[1],title='',ax=ax[1])
    meds=(frame(frame(cc([y,x])).median()))
    med1=float(meds.iloc[0])
    med2=float(meds.iloc[1])
    #print(med1,med2)
    meds.columns=['medians']
    meds.index=['-','-']
    res=frame(['Equal Medians',med1,med2,float(res[0]),float(res[1])],columns=['Mann-Whitney-Wilcoxon'],index=['H0:','Med-Y','Med-X','MWW-stat','P-Val'])   
    return res.T.round(6)

def m_mww_testa(df,y='y',x=0,plot=plotdefault):
    Y=df[y]
    try:
        X=df[x]
        X=X.dropna()
    except:
        print('invalid x variable')  
       
    Y=Y.dropna()
    
    return m_mww_test_(Y,X,plot)

def m_mww_test(df,y='y',x='',groupby='',group1='',group2='',plot=plotdefault):
    #print(df)
    if x == '':
        v1=df[df[groupby]==group1][y]
        v2=df[df[groupby]==group2][y]
        #print(v1,v2)
        out=m_mww_test_(v1,v2,plot)
    else:
        out=m_mww_testa(df,y=y,x=x,plot=plot)  
    return out

def np_test(df,y='y',x='',mu='',groupby='',group1='',group2='',plot=plotdefault):
    y,x,mu,groupby,group1,group2=warn(y=y,x=x,mu=mu,groupby=groupby,group1=group1,group2=group2)
    if mu=='':
        if x == '':
            v1=df[df[groupby].astype(str)==group1][y]
            v2=df[df[groupby].astype(str)==group2][y]
            #print(v1,v2)
            out=m_mww_test_(v1,v2,plot)
        else:
            out=m_mww_testa(df,y=y,x=x,plot=plot)
    else:
        print(' Since mu is active, a non-parametric Wilcoxon Test on the y variable is performed, other active cells ignored')
        out=m_w_test(df,y,x=float(mu),plot=plot)
    plt.show()    
    return out

#for equivalent variances (X and Y two variables)
def v_f_test_(Y,X,plot=plotdefault):
    Y,X=dropthenas(Y,X)
    if rows(Y)<2:
        print('not enough observations in Y','obs=',rows(Y)); stop()
    if rows(X)<2:
        print('not enough observations in X','obs=',rows(Y)); stop()
    s1=stdc(twodma(Y))
    s2=stdc(twodma(X)) 
    n1=rows(Y)
    n2=rows(X)
    F1 = float((s1/s2)**2)
    F2 = float((s2/s1)**2)
    pval1 = float(fdist.cdf(F1, n1-1, n2-1))
    if pval1>0.5:
        pval1=1-pval1
    pval2 = float(fdist.cdf(F2, n2-1, n1-1))
    if pval2>0.5:
        pval2=1-pval2
    out=frame((([['Common Var','Common Var'],[F1,F2],[pval1,pval2]])))
    
    out.columns=['Rat1/2','Rat2/1']
    out.index=['H0:','F-stat','P-val']
    #print(out)
    if plot==True:
       fig,ax=plt.subplots(ncols=2)
       plotbarpval(float(out.loc['P-val','Rat1/2']),title='HO: variances are the same',ax=ax[0]) 
       plotpiepval(float(out.loc['P-val','Rat1/2']),title='',ax=ax[1])
    #print(out)   
    return out.round(6).T

def v_f_testa(df,y='y',x=0,plot=plotdefault):
    Y=df[y].astype(float)
    try:
        X=df[x].astype(float)
    except:
        print('invalid x variable')    
    return v_f_test_(Y,X,plot)

#As above for pandas data frames which are stacked
def v_f_test(df,y='y',x='',groupby='',group1='',group2='',plot=plotdefault):
    if x == '':
        v1=df[df[groupby]==group1][y].astype(float)
        v2=df[df[groupby]==group2][y].astype(float)
        out=v_f_test_(v1,v2,plot)
    else:
        out=v_f_testa(df,y=y,x=x,plot=plot)
    return out

def v_testi(df,y='y',x='',mu='',groupby='',group1='',group2='',plot=plotdefault):
    y,x,mu,groupby,group1,group2=warn(y=y,x=x,mu=mu,groupby=groupby,group1=group1,group2=group2)
    if x == '':
        if mu=='':
            v1=df[df[groupby].astype(str)==group1][y].astype(float)
            v2=df[df[groupby].astype(str)==group2][y].astype(float)
            out=v_f_test_(v1,v2,plot)
        else:
            out=v_chi_test(df,y,x=float(mu),plot=plot)
    else:
        out=v_f_testa(df,y=y,x=x,plot=plot)
    plt.show()    
    return out


def m_t_test_(Y,X=0,paired=False,plot=plotdefault):
    title=''
    Y,X=dropthenas(Y,X)
    if paired:
       Y=twodma(Y)-twodma(X)
       X=0
      
    if rows(Y)<2:
        print('not enough observations in Y','obs=',rows(Y)); stop()
    '''
    if rows(X)<2:
        print('not enough observations in X','obs=',rows(X)); stop()
    '''    
    if stdc(twodma(X))==0:
        tst=ttest_1samp(Y,X)
        p1=float(tst[1])
        if paired==False:
            out=frame([float(meanc(Y)),float(X),float(tst[0]),float(tst[1])],index=['Mean','H0:Mean=','t-stat','P-val'],columns=['t-test']).T
        else:
            out=frame([float(meanc(Y)),float(X),float(tst[0]),float(tst[1])],index=['Mean(Y-X)','H0:MeanDiff=','t-stat','P-val'],columns=['t-test']).T
        if plot==True:
            fig,ax=plt.subplots(nrows=2)
            fig.suptitle('Test for Sample Mean =' + str(X))
            plotbarpval(p1,plot,ax=ax[0])
            plotpiepval(p1,plot,ax=ax[1])
            plotdist(tst[0],tail='s',distr=tdist,df=rows(Y)-1,value=True)
    else:    
        t1,p1=ttest_ind(Y,X,equal_var=True)
        t2,p2=ttest_ind(Y,X,equal_var=False)
        t1=float(t1)
        t2=float(t2)
        p1=float(p1)
        p2=float(p2)
        means=(frame(frame(cc([Y,X])).mean()))
        mx=float(means.iloc[1])
        my=float(means.iloc[0])
        out=frame([[my,mx,t1,p1],[my,mx,t2,p2]])
        out.columns=['Mean-y','Mean-x','t-stat','P-val']
        out.index=['equal var','unequal var']
        pval='p-val for H0: equal means'
        #out.index=['t-val',pval]
        #means=frame([float(meanc(Y)),float(meanc(X))]).T
        #means.index=['Means']
        #means.columns=['y','x']
        confidence=frame([[p1,1-p1],[p2,1-p2]])
        confidence.index=['eq var','uneq var']
        confidence.columns=['Sig','Conf']
        if plot==True:
            fig,ax=plt.subplots(nrows=2,ncols=2)
            ax[1,0].set_title='even'
            confidence.plot(kind='bar',ax=ax[0,0],rot=0,title='H0: Equal Means: ' + title,ylim=(-.1,1),grid=True,stacked=False,color=['red','green'])    
            plotpiepval(p1,title='',ax=ax[1,0])
            plotpiepval(p2,title='',ax=ax[1,1])
            #ax[0,1].table(0.5, 0.5, str(out.T.round(4)),
            #        verticalalignment='top', horizontalalignment='center',
            #        color='green', fontsize=12)
            #outshort=out.loc['equal var','unequal var']
            
            #ax[0,1].text(0.1,0.5,str(outshort.round(4)),fontsize=12)
            ax[0,1].set_axis_off()
            plotdist(t1,tail='s',distr=tdist,df=rows(Y)+rows(X)-2,value=True)
            #table(ax[0,1], confidence.round(3),  loc='center') 
            #fig.delaxes(ax[0,1])
            plt.show()
            
    return out.round(6)

def m_t_testa(df,y='y',x=0,paired=False,plot=plotdefault):
    try:
        Y=df[y].astype(float)
    except:
        pass
    try:
        X=df[x].astype(float)
    except:
        X=float(x)    
    return m_t_test_(Y,X,paired=paired,plot=plot)

def m_t_testi(df,y='y',x='',mu='',groupby='',group1='',group2='',paired=False,plot=plotdefault): 
        y,x,mu,groupby,group1,group2=warn(y=y,x=x,mu=mu,groupby=groupby,group1=group1,group2=group2)                      
        if x=='' and mu=='':
            v1=df[df[groupby].astype(str)==group1][y].astype(float)
            v2=df[df[groupby].astype(str)==group2][y].astype(float)
            v1=twodma(v1)
            v2=twodma(v2)
            out=m_t_test_(v1,v2,plot=plot)
        else:
            if mu=='':
                out=m_t_testa(df,y=y,x=x,paired=paired,plot=plot)
            else:
                paired=False
                out=m_t_testa(df,y=y,x=mu,paired=paired,plot=plot)
        return out

def m_t_test(df,y='y',x='',mu='',groupby='',group1='',group2='',paired=False,plot=plotdefault):
                
        if x=='' and mu=='':
            v1=df[df[groupby]==group1][y].astype(float)
            v2=df[df[groupby]==group2][y].astype(float)
            v1=twodma(v1)
            v2=twodma(v2)
            out=m_t_test_(v1,v2,plot=plot)
        else:
            if mu=='':
                out=m_t_testa(df,y=y,x=x,paired=paired,plot=plot)
            else:
                paired=False
                out=m_t_testa(df,y=y,x=mu,paired=paired,plot=plot)
        return out

def v_tests(data):
    test=v_testi
    u=['']+list(data.columns)
    return interact_manual(test,df=fixed(data),y=u,x=u,groupby=u)

def t_tests(data):
    test=m_t_testi
    u=['']+list(data.columns)
    return interact_manual(test,df=fixed(data),y=u,x=u,groupby=u)

def np_tests(data):
    test=np_test
    u=['']+list(data.columns)
    return interact_manual(test,df=fixed(data),y=u,x=u,groupby=u)


def contable_(mat,correction=False,plot=plotdefault,returnouts=True):
    if rows(mat)==1 or cols(mat)==1:
        if rows(mat)==1:
            mat=twodma(mat.T)
        otable=twodma(mat)
        dof=rows(otable)-1
        etable=meanc(otable)*ones(rows(otable))
        dtable=((mat-etable)**2)/etable
        chi=sumc(dtable)
        p=1-chidist(dof).cdf(chi)
        
    else:    
        chi,p,dof,etable=chi2_contingency(mat,correction)
        
    dtable=((mat-etable)**2)/etable
    
    otable=frame(mat)
    rowt=otable.index
    colt=otable.columns
    
    out=frame(['Independence',chi,p,dof])
    
    rsum=frame(otable.sum(axis=1))
    rsum.columns=['rsum']
    outs=cc([otable,rsum])
    csum=frame(outs.sum(axis=0))
    
    csum.columns=['csum']
    
    outs=cc([outs.T,csum]).T
    
    otable=outs
    
    etable=frame(etable)
    etable.columns=colt
    etable.index=rowt
    etable=cc([etable,rsum])
    etable=cc([etable.T,csum]).T
    
    
    out.index=['H0:','chi-square','P-val','deg-freedom']
    out.columns=[''] 
    outs={'summary':out.T,'observed':otable,'expected':etable,'dif':dtable}
    #print(out['summary'].round(6))
    print('Test\n',out.T)
    print('observed\n',otable)
    print('expected \n',etable)
    if plot==True:
       fig,ax=plt.subplots(ncols=2)
       plotbarpval(float(outs['summary']['P-val']),title='HO: rows and columns are independent',ax=ax[0]) 
       plotpiepval(float(outs['summary']['P-val']),title='Independence',ax=ax[1])
       plotdist(chi,tail='u',distr=chidist,df=dof,value=True)
       #plotpiepval(out['summary'].loc['p-val'])
    if returnouts:
        return outs

def categories(series):
    return range(int(series.min()), int(series.max()) + 1)

def contable(df,y='y',x='x',correction=False,plot=plotdefault,returnouts=True):
    df=df[[y,x]]
    df=df.dropna()
    df_col1,names1 = stringtonumber(df[y])
    df_col2,names2 = stringtonumber(df[x])
    
    result = [[sum((df_col1 == cat1) & (df_col2 == cat2))
               for cat2 in categories(df_col2)]
              for cat1 in categories(df_col1)]
    result=frame(result)
    result.index=names1
    result.columns=names2
    plt.show()
    return contable_(result,correction,plot,returnouts)

def contables(data):
    u=['']+list(data.columns)
    test=contable
    return interact_manual(test,df=fixed(data),y=u,x=u,returnouts=fixed(False))

def pivots(data,columns=None,index=None,values=None,aggfunc=('mean','median')):
    return pivot_table(data=data,columns=columns,index=index,values=values,aggfunc=aggfunc)

def pivot(data):
    u=[None]+list(data.columns)
    ag={'mean':'mean','median':'median','mean, median':('mean','median'),'count':'count','mean, stdv, sem':('mean','std','sem'),'max,min':('max','min')}
    return interact_manual(pivots,data=fixed(data),columns=u,index=u,values=u,aggfunc=(ag))
    #return interact_manual(pivot_,data=fixed(data),columns=u,index=u,values=u,aggfunc=fixed(('mean','median')))
    
def stringtonumber(series):
    x=list(set(series))
    x.sort()
    z=list()
    for i in range(rows(series)):
        for j in range(len(x)):
            if series.iloc[i]==x[j]:
                z=z+[j]
    return frame(z)[0],x

def converttorank(series,mapping):
    x=list()
    s=series.copy()
    k=0
    for i in series:
        for j in range(len(mapping)):
            if i==mapping[j]:
                s.iloc[k]=j+1
        k=k+1           
    return s

def recode(df,y,old,new):
    dfs=df.copy()
    for i in dfs.index:
        k=0
        for j in old:
            if dfs.loc[i,y]==j:
                dfs.loc[i,y]=new[k]
            k=k+1
    if type(dfs.loc[i,y])==float:
        dfs[y]=dfs[y].astype(float)
    if type(dfs.loc[i,y])==int:
        dfs[y]=dfs[y].astype(int)    
    return dfs    

def rename(df,oldname='',newname=''):
    dfc=df.copy()
    return dfc.rename(columns={oldname:newname})


def ktau_(y,x,plot=plotdefault):
    mat=cc([y,x])
    mat=frame(mat) 
    mat.columns=['y','x']
    mat=mat.dropna()
    out=sps.kendalltau(mat['y'],mat['x'])
    t=out[0]
    p=out[1]
    z=frame(['No Correlation',t,p], index=['H0:','Corr','P-val'],columns=['Kendall-Tau'])
    if plot==True: 
        fig,ax=plt.subplots(ncols=2)
        title= 'HO: No Correlation'
        plotbarpval(p,title=title,ax=ax[0]) 
        plotpiepval(p,title='',ax=ax[1])
    return z.T
    
def ktau(data,y='y',x='x',rnky='',rnkx='',plot=plotdefault):
    if rnkx !='':
        xs=converttorank(data[x],rnkx)
    else:
        xs=data[x]    
    if rnky !='':
        ys=converttorank(data[y],rnky)
    else:
        ys=data[y]   
    
    mat=cc([ys,xs])
    mat=mat.dropna()
    z=ktau_(mat[y],mat[x],plot)
    return z 

def corr_(y,x,spearman=False,plot=plotdefault):
        mat=cc([y,x])
        mat=frame(mat) 
        mat.columns=['y','x']
        mat=mat.dropna()
        if spearman==False:
            out=sps.pearsonr(mat['y'],mat['x'])
        else:
            out=sps.spearmanr(mat['y'],mat['x'])
        
        t=out[0]
        p=out[1]
        if spearman==False:
                z=frame(['No Correlation',t,p], index=['H0:','Corr','P-val'],columns=['Pearson Correlation'])
        else:
                z=frame(['No Correlation',t,p], index=['H0:','Corr','P-val'],columns=['Spearman Correlation'])
      
        if plot==True:
            fig,ax=plt.subplots(ncols=2)
            title= 'HO: No Correlation'
            plotbarpval(p,title=title,ax=ax[0]) 
            plotpiepval(p,title='',ax=ax[1])
            
        return z.round(6).T 

def corr(data,y='y',x='x',rnky='',rnkx='',spearman=False,kendtau=False,plot=plotdefault): 
    if kendtau==False:
        if rnkx !='':
            xs=converttorank(data[x],rnkx)
        else:
            xs=data[x]    
        if rnky !='':
            ys=converttorank(data[y],rnky)
        else:
            ys=data[y]
    
        mat=cc([ys,xs])    
        mat=mat.dropna()
        #print(mat)
        z=corr_(mat[y],mat[x],spearman,plot)
    else:
        z=ktau(data,y=y,x=x,rnky=rnky,rnkx=rnkx,plot=plot)
    plt.show()
    return z.round(6) 

def corrs(data,y='y',x='x',rnky='',rnkx='',spearman=False,kendtau=False,plot=plotdefault): 
    u=['']+list(data.columns)
    interact_manual(corr,data=fixed(data),y=u,x=u,rnky=fixed(''),rnkx=fixed(''))
    return
          
def mean_ci(df,y='',groupby='',confidence=0.95,):
    df=frame(df)
    if y !='' and groupby !='':
       df=df[[y,groupby]]
    elif y != '':
       df=df[y] 
       df.columns=y
    dist=tdist._ppf((1+confidence)/2., rows(df)-1)
    if groupby == '':
        dfm=df.mean()
        dfs=df.sem()
    else:
        dfm,dfs=groupby_(df,groupby)  
    dfsu=dfm+dfs*dist
    dfsl=dfm-dfs*dist   
    if rows(dfm)==1:
        dfs=frame([dfm,dfsl,dfsu,dfs]).T
    else:    
        dfs=cc([dfm,dfsl,dfsu,dfs])    
    dfs.columns=['mean','lower-CI','upper-CI','SE']
    return dfs

def mean_CIs(data,y='',groupby='',confidence=0.95,plot=False):
    u=['']+list(data.columns)
    return interact(mean_cis,data=fixed(data),groupby=u,y=u,confidence=(.5,.99,.01))    

def mean_cis(data,y='',groupby='',confidence=0.95,plot=False):
    if plot:
        return plotmeanbar(data,y=y,groupby=groupby,confidence=confidence) 
    else:
        return mean_ci(data,y=y,groupby=groupby,confidence=confidence)     

def sumc(x):
    x=twodma(x)
    return twodm(np.sum(x,axis=0))

def meanc(x):
    x=twodma(x)
    return twodm(np.mean(x,axis=0))

def medianc(x):
    x=twodma(x)
    return twodm(np.median(x,axis=0))
    
def stdc(x):
    x=twodma(x)
    return twodm(np.std(x,axis=0))
    
def minc(x):
    x=twodma(x)
    return twodm(np.min(x,axis=0))    

def maxc(x):
    x=twodma(x)
    return twodm(np.max(x,axis=0))    

def cc(ob):
    if type(ob[0])==pd.core.frame.DataFrame or type(ob[0])==pd.core.series.Series:
        return pd.concat(ob,axis=1)
    else:
        return concatenate(ob,axis=1)

def rc(ob):
    if type(ob[0])==pd.core.frame.DataFrame or type(ob[0])==pd.core.series.Series:
        return pd.concat(ob,axis=0)
    else:
        return concatenate(ob,axis=0)
    
def reshape(x,rows,cols):
    return np.reshape(x,[cols,rows]).T

def mode_(v,y=0,bins=10):
    try:
        x=frame(v)
        x=x[y].dropna()
        c=x.value_counts()
        count=(c >1).sum()
        flag=False
        if count >0:
            mod=c.idxmax()
            flag=True
        else:
            bins=bins
            
            if rows(x)>50:
                bins= 10+ round(rows(x)/50)
                bins=bins.astype(int)
                
            res=np.histogram(x,bins=bins)
            mod=(res[1][res[0].argmax()]+res[1][res[0].argmax()+1])/2
    except:
        mod=np.nan
    return mod

def mode(dat,y=''):
        if y=='' and type(dat)==pd.core.frame.DataFrame:
           y=list(dat.columns)
        if type(y) != list:
           y=[y]
        dats=dat.copy()
        rec=list()
        for i in y:
            m=mode_(dats[i],i)
            rec=rec+[m]
            
        return frame(rec,index=y,columns=['mode'])  
    
def describe(data,y=''):
    
    if y=='' and type(data)==pd.core.frame.DataFrame:
        y=list(data.columns)
        #print(y)
    if type(y) != list:
        #print(type(y))
        y=[y]
        dat=data[y]
    else:
        dat=data[y]
    for i in dat.columns:
        try: 
            dat[i]=dat[i].astype(float)
        except:
            pass
    
    valid,same,h,n=datatypes_(dat)
    
    if valid==False:
        print('Since, the data frame includes a variable that has more than one type, you should remove it or alter the variable, see below')
        stop()
    elif same==False:
        print('The data frame includes numeric and non-numeric variables, do one or the other only, see below ')
        tp=datatypes(dat)
        stop()
       
    z=dat.describe()
    z.loc['kurt']=dat.kurt()
    z.loc['skew']=dat.skew()
    z.loc['sem']=dat.sem()
    #print(mode(dat).T)
    z=pd.concat([z,mode(dat).T],axis=0)
    z=z.dropna()
    return z

def datatypes__(d):
    d=frame(d)
    h=list()
    for i in range(rows(d)):
        x=d.values[i,0]
        try:
            z=np.isnan(x)
        except:
            z=False
        if ~z:
            h=h+[type(x)]
    s=list(set(h))
    return s

def datatypes_(d):
    valid=True
    same=True
    d=frame(d)
    h=list()
    n=list()
    j=1
    for i in d.columns:
        dt=datatypes__(d[i])
        if j>1:
           if dt !=dt0:
              same=False
        h=h+[dt]
        n=n+[int(len(dt))]
        if int(len(dt))>1:
           valid=False  
           print('Warining column', i, 'has more than one type data')
        dt0=dt
        j=j+1
    return valid,same,h,n

def datatypes(d,checkit=True):
    check=False
    if checkit:
        check=datatypes_(d)
    j=0
    for i in d.columns:
        v=type(d[i].loc[1])
        n=len(inquire(d,i))
        if n<10:
           n=inquire(d,i) 
        else:
           n=['number of unique values= ', len(inquire(d,i))]
        if v==np.int64:
           v=['scale integer:',n]
        if v==np.int32:
           v=['scale integer:',n]
        if v==np.float64:
           v=['scale continuous:',n]
        if v==np.float32:
           v=['scale continuous:',n] 
        if v==str:
           v=['categorical, string:',n]
        if v==np.int:
           v=['integer',n]
        print(i,':',v)
        j=j+1
    return check  

def variableviewf(d,checkit=True,frameit=True):
    di=dict()
    check=False
    if checkit:
        check=datatypes_(d)
    j=0
    for i in d.columns:
        v=type(d[i].loc[1])
        try:
            n=len(inquire(d,i))
        except:
            n=''
        if n<10 and n !='':
           ni=inquire(d,i) 
        else:
           ni=['number of unique values= ', n]
        if v==np.int64:
           v=['scale integer:',ni]
        elif v==np.int32:
           v=['scale integer:',ni]
        elif v==np.float64:
           v=['scale continuous:',ni]
        elif v==np.float32:
           v=['scale continuous:',ni] 
        elif v==str:
           v=['categorical, string:',ni]
        elif v==np.int:
           v=['integer',ni]
        else:
           v=[v,ni] 
        if frameit:
            if j==0:
                di={i:v}
            else:
                di.update({i:v})
        else:        
            #print(i,':',v)
            pass
        j=j+1
    if frameit:        
        df=frame(di.values(),index=di.keys())
        df.columns=['Type','Unique Vals']
    else:
        df=''
    return df

def inquire(data,var):
    if var !='':
        return set(data[var])
    else:
        return ''
    
    
def variableview(data,interact=True):
    if interact:
        datatypes(data,checkit=False)
        u=['']+list(data.columns)
        return interact_manual(inquire,data=fixed(data),var=u)
    else:
        return variableviewf(data)
    

def annuity(x,r,t):
    return (x/r)*(1-(1+r)**-t)  


def from1(df,start=1):
    dfc=df.copy()
    dfc.index=(np.arange(start,rows(df)+start))
    return dfc

def findex(x,start=1):
    return from1(frame(x),start)

def proportion(df,y,by):
    s=df[y]==by
    return meanc(s)

def proportioni(df,y,by):
    s=df[y].astype(str)==by
    return meanc(s)


def plotbarpval(p,title='H0',ax=None):
    val=round(p,4)
    f=frame([p,1-p]).T
    f.columns=['p-value =' +str(val) ,'confidence to reject H0\n=' +str(1.0-val)]
    f.plot(kind='bar',rot=0,title=title,ylim=(-.1,1),grid=True,stacked=False,color=['red','green'],ax=ax)
    return

def plotpiepval(p,title='',ax=None):
    p=round(p,4)
    f=frame([p,1-p])
    f.index=['p-value','confidence \nto reject H0']
    f.columns=[title]
    f.plot.pie(y=title,colors=['red','green'],fontsize=12,autopct='%.2f',ax=ax,legend=False)
    return 
    
def plotpie(df,y=0,ax=None):
    #Needs to be a series or frame where the indexes are the slices
    df=frame(df)
    df.plot.pie(y=y,autopct='%.2f',ax=ax)
    return

def plotbar(df,ax=None):
    #Needs to be a series or frame where the indexes are the slices
    df=frame(df).T
    df.plot(kind='bar',grid=True)
    return

def plotsequence(data,y=''):
    data=frame(data)
    if y !='':
        data[y].plot(grid=True)
    else:
        data.plot(grid=True)

def plothist(df,y=0,bins=50,normed=False,normal=False,groupby='',group=''):
    
    fig,ax=plt.subplots(nrows=1)
    dfs=df[y].copy()
    if groupby !='':
        s=df[groupby].astype(str)==group 
        dfs=dfs[s]
    dfs=dfs.dropna()
    if normal:
        normed=True
        dfs=frame(dfs)
        minx=float(minc(dfs[y]))
        maxx=float(maxc(dfs[y]))
        mean=meanc(dfs[y])
        stdv=stdc(dfs[y])
        #print(minx,maxx,mean,stdv)
        xs=np.linspace(minx,maxx,1000)
        ys=twodmca(norm(mean,stdv).pdf(xs))
        #print(ys)
        ys=frame(ys,index=xs,columns=['Normal'])
        ys.plot(ax=ax,title='Distribution')
    dfs.plot(kind='hist',y=y,bins=bins,density=normed,grid=True,ax=ax)
    plt.show()
    return

def histogram(data):
    u=['']+list(data.columns)
    return interact_manual(plothist,df=fixed(data),y=u,bins=(10,100,1),groupby=u)

def plotscatter(df='',y=0,x=0,line=False):
    s=type(df)==pd.core.frame.DataFrame
    if s:
       df.plot(kind='scatter',y=y,x=x,grid=True)
       X=twodma(df[x])
       Y=twodma(df[y])
       if line:
           w = np.linalg.lstsq(np.hstack((X, np.ones((len(X),1)))), Y)[0]
           xx = np.linspace(*plt.gca().get_xlim()).T
           # plot best-fit line
           plt.plot(xx, w[0]*xx + w[1], '-k')
    else:    
        df=frame(cc([df,y])) 
        y=df.columns[0]
        x=df.columns[1]
        df.plot(kind='scatter',y=y,x=x,grid=True)
    plt.show()
    return

def scatterplot(data):
    u=['']+list(data.columns)
    return interact_manual(plotscatter,df=fixed(data),y=u,x=u)


def plotcounts(df='',y=0,normed=False,pie=False,confidence=0.95,plot=False):
    s=type(y)==pd.core.frame.DataFrame
    t=type(y)==pd.core.series.Series
    u=type(df)==pd.core.series.Series
    if type(df)==np.ndarray:
         df=frame(df)
    if t:
       z=y 
    elif u:
       z=df
    else:
       z=df[y]
    z=z.dropna()
    #z=frame(array(z))
    #print(z)
    x=z.value_counts()
    zval=norm(0,1).ppf((1+confidence)/2.0)
    
    if normed==True or pie==True:
        x=x/len(z)
        stdv=sqrt(x*(1-x)/len(z))
    x=frame(x)
    x=x.sort_index(axis=0)
    if plot:
        if pie==False:
            if normed==True:
                x.plot(kind='bar',yerr=zval*stdv,capsize=20,grid=True)
            else: 
                x.plot(kind='bar',grid=True)
        else:
            plotpie(x)
        plt.show()    
    return frame(x)

def counts(data):
     u=['']+list(data.columns)
     return interact_manual(plotcounts,df=fixed(data),y=u,confidence=(0.5,.99,.01))

def plotdist(val,tail='u',distr=norm,mean=0,stdv=1,df=10,df2=5,ub=1,lb=0,value=False):
    if distr==uniform:
        dist=uniform(lb,ub-lb)
    if distr==beta:
        dist=beta(df,df2)
    if distr==norm:
        dist=norm(mean,stdv)
    if distr==lognorm:
        mu,sigma=reverselognorm(mean,stdv)
        dist=lognorm(loc=0,scale=exp(mu),s=sigma)
    if distr==tdist:
        dist=tdist(df)
    if distr==chidist:
        dist=chidist(df)
    if distr==fdist:
        dist=fdist(df,df2)    
        
    
        
    val=twodma(val)    
    if tail[0]=='s' and value==False:
        val=[val/2,val/2]
    if tail[0]=='s' and value==True and len(val)==1:
        val=[-abs(val),abs(val)]
        
    val=twodma(val)
    
    if shape(val)[0] > 1:
       tail='s' 
    #print(types(val))
    if value==True:
       if tail[0]=='u': 
           val=1-dist.cdf(val) 
       elif tail[0]=='l':
           val=dist.cdf(val)
       elif tail[0]=='s':
           #print(types(dist.cdf(float(val[0]),mean,stdv)))
           v0=dist.cdf(float(val[0])) 
           v1=1-dist.cdf(float(val[1]))
           val=[v0,v1]
    val=twodma(val)
    
    
    if tail[0]=='u':
        tail='upper'
    if tail[0]=='l':
        tail='lower'
    if tail[0]=='s' or tail[0]=='b':
        tail='split'    
    
    if distr==norm or distr==uniform: 
        span=(dist.isf(0.99995),dist.isf(0.00005))
    
    if distr==beta:
        print('yes')
        span=(dist.isf(0.99),dist.isf(0.01))
        
    if distr==tdist:
        if df>5:
            span=(dist.isf(0.99995),dist.isf(0.00005))
        elif df>1:
            span=(dist.isf(0.995),dist.isf(0.005))
        elif df==1:
            span=(dist.isf(0.99),dist.isf(0.01))    
    if distr==lognorm or distr ==chidist or distr==fdist:
        span=(0,dist.isf(0.01))
       
    x=np.linspace(span[0],span[1],1000)
    
    val_=val 
    
    if tail[0]=='u':
        cvu=dist.isf(val[0])
        cvl=0
    elif tail[0]=='l':
        cvl=dist.isf(1.0-val[0])
        cvu=0
    elif tail[0]=='s':
        cvl=dist.isf(1-val[0])
        cvu=dist.isf(val[1])
      
        if shape(val)[0]>1:
            val_=val[0]+val[1]
    cvl=squeeze(cvl)
    cvu=squeeze(cvu)
    #print(cvl,cvu,tail)
    #stop()
    if distr==norm:
        suptitle='PDF Normal with mean ' + str(mean) +  ' and standard deviation '+ str(stdv) 
    elif distr==lognorm:
        suptitle='PDF Log Normal with mean ' + str(mean) +  ' and standard deviation '+ str(stdv) 
    elif distr==uniform:
        suptitle='PDF Uniform'
    elif distr==beta:
        suptitle='PDF Beta'    
    elif distr==tdist and df>=10000:
        suptitle='PDF Standard Normal'
    elif distr==tdist:
        suptitle='PDF Student-T, ' + str(df) + ' degrees of Freedom'
    elif distr==fdist:
        suptitle='PDF F Distribution, ' + str(df) +','+ str(df2)+ ' degrees of Freedom'    
    elif distr==chidist:
         suptitle='PDF Chi Square, ' + str(df) + ' degrees of Freedom'
    y=dist.pdf(x)
    z=dist.cdf(x)
    
    val_=squeeze(val_)
    sgn=1
    if tail[0]=='s':
        label='RED' + ' prob='+ str(val_)[0:5]+'\nGREEN' + ' prob='+ str(1-val_)[0:6]+ '\n Vals=' + str(round(cvl,6))[0:6] + ','+str(round(cvu,5))[0:5]  
    elif tail[0]=='u':
        label=tail.upper()+ ' RED prob='+ str(val_)[0:5]+ '\n Vals=' +str(round(cvu,6))[0:5]
    elif tail[0]=='l' :   
        label=tail.upper()+ ' RED prob='+ str(val_)[0:5]+ '\n Vals=' + str(round(cvl,6))[0:5]
        
    fig,ax =plt.subplots(nrows=2)    
    dfs=frame(y,index=x,columns=[label])
    dft=frame(z,index=x,columns=['CDF'])
    if maxc(y)<1000:
        dfs.plot(ylim=(0.001,maxc(y)+.01),title=suptitle,grid=True,ax=ax[0])
    else:    
        dfs.plot(title=suptitle,grid=True,ax=ax[0])
    dft.plot(ylim=(0.001,1),title='CDF',grid=True,ax=ax[1])
    section1=np.linspace(cvu,span[1],1000)
    section2=np.linspace(span[0], cvl, 1000)
    
    if tail[0]=='b' or tail[0]=='s':
        section=rc([section1,section2])
    elif tail[0]=='l':
        section=section2
    elif tail[0]=='u':
        section=section1     
    
    ax[0].fill_between(x,dist.pdf(x),color='green')
    ax[0].fill_between(section,dist.pdf(section),color='red')
    
    ax[1].fill_between(x,dist.cdf(x),color='blue')
    #ax[1].fill_between(section,dist.cdf(section),color='blue')
    plt.show()
    return


def groupby_(df,group):
    dfm=df.groupby(group).mean()
    dfs=df.groupby(group).sem()
    del dfm.columns.name
    del dfs.columns.name
    del dfm.index.name
    del dfs.index.name
    return dfm.iloc[:,0],dfs.iloc[:,0]


def plotmeanbar(df,confidence=0.95,y='',groupby='',table=False,gap=0.1,title='means'):
       
    if y != '' and groupby !='':
        df=df[[y,groupby]]
    elif y != '':
       df=df[y]
    if type(df) !=pd.core.frame.DataFrame:
       df=frame(df)
    
    dist=tdist._ppf((1+confidence)/2., rows(df)-1)
    if groupby == '':
        dfm=df.mean()
        dfs=df.sem()
    else:
        dfm,dfs=groupby_(df,groupby)
    #dfm=frame(dfm)
    #dfs=frame(dfs)
    dfs=dfs*dist
    fig, ax= plt.subplots()
    fig.suptitle(title)
    ax.set_title('')
    dfm.columns=['Score']
    ax=dfm.plot(kind='bar',yerr=dfs,capsize=20, color='grey',alpha=0.5,legend='',grid=True,rot=0,ax=ax,fontsize=12)

    for i, label in enumerate(list(dfm.index)):
        score = dfm.iloc[i].round(4)
        ax.annotate(str(score), (i, score + gap))
    
    dfs=mean_ci(df, confidence=confidence,y=y,groupby=groupby)
    if table:
        fig.suptitle(str(dfs))
    plt.show()    
    return dfs

def plotmeanbars(df,confidence=0.95,y='',groupby='',transpose=False,colors=['grey','black','blue'],title='Means',alpha=0.75):
    dfs=df.copy()
    if y !='':
        dfs=cc([dfs[y],dfs[groupby]])
    dist=tdist._ppf((1+confidence)/2., rows(df)-1)
    mu=dfs.groupby(groupby).mean()
    cis=dist*dfs.groupby(groupby).sem()
    if transpose:
       mu=mu.T
       cis=cis.T         
    mu.plot(kind='bar',yerr=cis,capsize=10,legend=True,title=title,grid=True,colors=colors,alpha=alpha,fontsize=12 )
    return mu,cis


def plot_mean_CIs(df,y='',groupby='',confidence=0.95):
    u=['']+list(df.columns)
    return interact(plotmeanbar,df=fixed(df),groupby=u,y=u,confidence=(.5,.99,.001),gap=(0,10,.01),table=fixed(False),title=fixed('Means and Confidence Intervals'))    



def plotbox(df,y='',groupby=''):
    fig, ax= plt.subplots()
    if y != '' and groupby !='':
        df=df[[y,groupby]]
    elif y != '':
       df=df[y]
    if type(df) !=pd.core.frame.DataFrame:
       df=frame(df)
    
    if groupby=='':
        df.boxplot(grid=True,ax=ax)
        plt.show()
    else:
        df.boxplot(by=groupby,grid=True,ax=ax)
        plt.show()
    return

def boxplot(data):
    u=['']+list(data.columns)
    return interact_manual(plotbox,df=fixed(data),y=u,groupby=u)
    
def color_negative_red(val):
    #print(val.iloc())
    """
    Takes a scalar and returns a string with
    the css property `'color: red'` for negative
    strings, black otherwise.
    """
    color = 'red'
    return 'color: %s' % color           
              
def presentval(y,r,plot=False,highlight=False):
    amounts=y.copy()
    amounts=frame(amounts)
    try:
        amounts.columns=['Amount']
    except:
        amounts.index=['Amount']
    n=rows(amounts)
    d=[1]
    for t in range(1,n):
        d=d+[(1+r)**-t]
    amounts['d-factor']=d
    amounts['Present Val']=amounts['d-factor']*amounts['Amount'] 
    amounts['Discounted Sum']=amounts['Present Val'].cumsum()
    amounts['Sum']= amounts['Amount'].cumsum()
    if plot==True:
        amounts[['Discounted Sum','Sum']].plot(grid=True)
        plt.show()
    y=amounts
    if highlight==True:
        y=y.style.applymap(color_negative_red,subset=pd.IndexSlice[rows(y)-1,['Discounted Sum']])          
    return y    

def plotuniform(n):
    fig,ax=plt.subplots(nrows=2)
    x=np.arange(1,n+1)
    
    ps=(x**0)/n
    cs=cumsumc(ps)
    f=frame(ps, columns=['x'])
    g=frame(cs, columns=['x'])
    f.index=x
    g.index=x
    f.plot(kind='bar',title='PMF Discrete Uniform n=' +str (n) ,ax=ax[0],grid=True,width=.1,color='green')
    g.plot(kind='bar',title='CDF Discrete Uniform n=' +str (n) ,ax=ax[1],grid=True,width=.1,color='blue')
    plt.xlabel('number')
    plt.ylabel('probability')
    addvalstobar(f,ax[0])
    addvalstobar(g,ax[1])
    plt.show()
    return

def plotbinom(n,p):
    fig,ax=plt.subplots(nrows=2)
    dist=binom(n,p)
    ps=dist.pmf(np.arange(0,n+1))
    cs=dist.cdf(np.arange(0,n+1))
    f=frame(ps, columns=['x successes'])
   
    g=frame(cs, columns=['x successes'])
    
    f.plot(kind='bar',title='PMF Binomial distribution with trials =' + str(n) + ' and probability= ' + str(p) ,ax=ax[0],grid=True,width=.1,color='green')
    g.plot(kind='bar',title='CDF Binomial distribution with trials =' + str(n) + ' and probability= ' + str(p) ,ax=ax[1],grid=True,width=.1,color='blue')
    plt.xlabel('number of successes in n trials')
    plt.ylabel('probability')
    ax[0].set_ylim(0,maxc(f)+.1)
    addvalstobar(f,ax[0])
    addvalstobar(g,ax[1])
    plt.show()
    return


def plotpoisson(mean):
    fig,ax=plt.subplots(nrows=2)
    dist=poisson(mean)
    n=round(dist.isf(.01))
    ps=dist.pmf(np.arange(0,n+1))
    cs=dist.cdf(np.arange(0,n+1))
    f=frame(ps, columns=['x successes'])
    g=frame(cs, columns=['x successes'])
    f.plot(kind='bar',title='PMF Poisson distribution with Mean =' + str(mean) ,ax=ax[0],grid=True,width=.1,color='green')
    g.plot(kind='bar',title='CDF Poisson distribution with Mean =' + str(mean) ,ax=ax[1],grid=True,width=.1,color='blue')
    plt.xlabel('number of events')
    plt.ylabel('probability')
    ax[0].set_ylim(0,maxc(f)+.1)
    addvalstobar(f,ax[0])
    addvalstobar(g,ax[1])
    plt.show()
    return

def plotnbinom(x,p,ntrials=False):
    fig,ax=plt.subplots(nrows=2)
    # x=Number of required successes
    # p=Probability of succeses
    dist=nbinom(x,p)
    n=round(dist.isf(0.01))
    ps=dist.pmf(np.arange(0,n+1))
    cs=dist.cdf(np.arange(0,n+1))
    f=frame(ps, columns=['x failure'])
    g=frame(cs, columns=['x failures'])
    if ntrials:
        f.index=f.index+x
        g.index=g.index+x
    f.plot(kind='bar',title='PMF Negative Binomial distribution with Number of Successes =' + str(x) + ' and probability= ' + str(p) ,ax=ax[0],grid=True,width=.1,color='green')
    g.plot(kind='bar',title='CDF Negative Binomial distribution with Number of Successes =' + str(x) + ' and probability= ' + str(p) ,ax=ax[1],grid=True,width=.1,color='blue')
    plt.xlabel('number of failures to get required number of successes')
    plt.ylabel('probability')
    ax[0].set_ylim(0,maxc(f)+.1)
    addvalstobar(f,ax[0])
    addvalstobar(g,ax[1])
    plt.show() 
    return

def addvalstobar(df,ax):
    for i, label in enumerate(list(df.index)):
            score = df.iloc[i].round(4)
            ax.annotate(str(float(score)), (i, score + 0.01))
    return   

def distmap(dist,x=0,):
    if 'binom' in str(dist.dist):
        if 'nbinom' in str(dist.dist):
            distr='NegBinomial' +str(dist.args)
            h=' failures'
        else:
            distr='Binomial'  +str(dist.args)
            h=' successes'
    elif 'poisson' in str(dist.dist):
        distr='Poisson' +str(dist.args)
        h=' events '
        
    z=str(x)         
    f=frame([dist.pmf(x),dist.cdf(x),dist.cdf(x-1),dist.sf(x),dist.sf(x-1)])
    f.index=['='+z+h ,'<=' +z+h,'<' +z+h,'>' +z+h,'>='+z+h]
    f.columns=['Probs']
    print(f)
    return f

def distcmap(dist,lower='',upper=''):
    if upper=='':
        p=dist.cdf(lower)
        return ['prob(x<=' + str(lower)+')',p],['prob x>='+ str(lower)+')',1-p]
    else:
        pu=dist.cdf(upper)
        pl=dist.cdf(lower)
        return ['prob(x<=' + str(lower)+')',pl],['prob(' + str(lower)+ '<=x<=' + str(upper)+')',pu-pl],['prob(x>='+str(upper)+')',1-pu]
                                                       
    

def plotlinefrompoints(point0,point1,lx='x',ly='y'):
    x0=point0[0]
    y0=point0[1]
    x1=point1[0]
    y1=point1[1]
    df=frame([[x0,y0],[x1,y1]],columns=[lx,ly])
    intercept,slope=solveline(x0,y0,x1,y1)
    if x0>x1:
       span=[x1-5,x0+5]
    else:   
       span=[x0-5,x1+5] 
    fig,ax=plt.subplots(nrows=1)
    fig.suptitle('',fontsize=20)
    eq1=textder(intercept,slope,1,0,0,x=lx,y=ly)
    xs=np.linspace(span[0],span[1],1000)
    line1=intercept+mtransform(xs,slope,1)
    line1=frame(line1,index=xs,columns=[eq1])
    line1.plot(title= eq1 + ':  through points (' +lx +',' +ly+') ='+ str(point0) + ' and ' + str(point1),grid=True,linewidth=2,ax=ax)
    df.plot(kind='scatter',y=ly,x=lx,ax=ax,grid=True,fontsize=18)
    ax.plot(x0,y0,'ob')
    ax.plot(x1,y1,'ob')
    ax.vlines(x0,0,y0, linestyles="dashed")
    ax.vlines(x1,0,y1, linestyles="dashed")
    ax.hlines(y1,0,x1, linestyles="dashed")
    ax.hlines(y0,0,x0, linestyles="dashed")
    ax.set_xlabel(lx,fontsize=14)
    ax.set_ylabel(ly,fontsize=14)
    plt.show()
    return



def plotline(a0,a1,span=(-1,1),lx='x',ly='y'):
    fig,ax=plt.subplots(nrows=1)
    eq1=textder(a0,a1,1,0,0,x=lx,y=ly)
    xs=np.linspace(span[0],span[1],1000)
    line1=a0+mtransform(xs,a1,1)
    line1=frame(line1,index=xs,columns=[eq1])
    line1.plot(title= eq1,grid=True,linewidth=2,ax=ax,fontsize=20)
    ax.set_xlabel(lx,fontsize=14)
    ax.set_ylabel(ly,fontsize=14)
    plt.show()
    return 

def plotbreakeven_exam(point0,point1,mr,addfc=0,addmc=0,addmr=0):
    x0=point0[0]
    y0=point0[1]
    x1=point1[0]
    y1=point1[1]
    fc,mc=solveline(x0,y0,x1,y1)
    
    sy0,sx0=plotbreakeven(fc,mc,mr)
    sy1=sy0
    sx1=sx0
    if mr>mc:
        profit2=-(fc+addfc)-(mc+addmc)*sx0 + (mr+addmr)*sx0
    else:
        profit2=-fc-addfc
        
    if addfc !=0 or addmr !=0 or addmc !=0:
        try:
            sy1,sx1=plotbreakeven(fc+addfc,mc+addmc,mr+addmr)
            if sx1>0:
                profit1=-fc-mc*sx1 + mr*sx1
            else:
                profit1=-fc
        except:
                profit1=-fc
    else:
        if mr>mc:
               profit1=0
        else:
               profit1=-fc
            

 
    return [fc,mc,mr], [fc+addfc,mc+addmc,mr+addmr],[sy0,sx0],[sy1,sx1],[profit1,profit2]

def plotbreakeven(fc,mc,mr,lx='x'):
    a0=fc
    a1=mc
    b0=0
    b1=mr
    A=twodma([[1,-mc],[1,-mr]])
    C=twodma([fc,0])
    try:
        B=inv(A).dot(C)
        solx=float(B[1])
        sol=True
    except:
        solx=0
        B=[' There is no solution','']
        sol=False
    if solx>0:
        bg=0
    else:
        bg=solx
    fig,ax=plt.subplots(nrows=1)
    eq1=textder(a0,a1,1,0,0,y='C',x=lx)
    eq2=textder(b0,b1,1,0,0,y='R',x=lx)
    xs=np.linspace(bg,solx+5,1000)
    line1=a0+mtransform(xs,a1,1)
    line2=b0+mtransform(xs,b1,1)
    line1=frame(line1,index=xs,columns=[eq1])
    line2=frame(line2,index=xs,columns=[eq2])
    if sol==True:
        sy=round(B[0],3)
        sx=round(B[1],3)
        line1.plot(title= eq1 + ' and  ' + eq2 + '\n Solution:'+ 'Cost=Revenue' +str(sy)+ ' Output=x=' + str(sx),grid=True,linewidth=2,ax=ax)
    else:
        sy=B[0]
        sx=B[1]
        line1.plot(title= eq1 + ' and  ' + eq2 + ':'+ sy + str(sx),grid=True,linewidth=2,ax=ax)
    line2.plot(grid=True,linewidth=2,ax=ax)
    ax.set_xlabel(lx,fontsize=14)
    ax.set_ylabel('C,R',fontsize=14)
    ax.plot(solx,sy,'ob')
    ax.vlines(solx,0,sy, linestyles="dashed")
    ax.hlines(sy,0,solx, linestyles="dashed")
    
    plt.show()
    return sy,sx


def plotSD(a0,a1,b0,b1,loglinear=False):
    #a) parameters supply, b) Demand 
    a0s=a0; a1s=a1; b0s=b0; b1s=b1;
    a0=-a0/a1
    a1=1/a1
    b0=-b0/b1
    b1=1/b1
    A=twodma([[1,-a1],[1,-b1]])
    C=twodma([a0,b0])
    B=inv(A).dot(C)
    solx=float(B[1])
    
    fig,ax=plt.subplots(nrows=1)
    if loglinear:
        eq1=textder(a0s,a1s,1,0,0, x='lnP',y='lnS')
        eq2=textder(b0s,b1s,1,0,0,   x='lnP',y='lnD')
    else:
        eq1=textder(a0s,a1s,1,0,0, x='P',y='S')
        eq2=textder(b0s,b1s,1,0,0,   x='P',y='D')
    xs=np.linspace(solx-5,solx+5,1000)
    if loglinear:
        xs=np.linspace(solx-1,solx+1,1000)
    line1=a0+mtransform(xs,a1,1)
    line2=b0+mtransform(xs,b1,1)
    
    if loglinear:
        line1=exp(line1)
        line2=exp(line2)
        xs=exp(xs)    
    
    line1=frame(line1,index=xs,columns=[eq1])
    line2=frame(line2,index=xs,columns=[eq2])
    
    sy=round(B[0],3)
    sx=round(B[1],3)
    
    if loglinear:
        sy=exp(sy)
        sx=exp(sx)
        
    line1.plot(title= eq1 + ' and  ' + eq2 + '\n Solution:'+ 'P=' +str(sy)+ '  Q=' + str(sx),grid=True,linewidth=2,ax=ax)
    line2.plot(grid=True,linewidth=2,ax=ax)
    ax.set_xlabel('Q',fontsize=14)
    ax.set_ylabel('P',fontsize=14)
    ax.plot(sx,sy,'ob')
    ax.vlines(sx,0,sy, linestyles="dashed")
    ax.hlines(sy,0,sx, linestyles="dashed")
    
    plt.show()
    return
    
def plotsimult(a0,a1,b0,b1,lx='x',ly='y'):
    A=twodma([[1,-a1],[1,-b1]])
    C=twodma([a0,b0])
    try:
        B=inv(A).dot(C)
        solx=float(B[1])
        sol=True
    except:
        B=[' There is no Solution','']
        solx=0
        sol=False
        
    fig,ax=plt.subplots(nrows=1)
    eq1=textder(a0,a1,1,0,0,x=lx,y=ly)
    eq2=textder(b0,b1,1,0,0,x=lx,y=ly)
    xs=np.linspace(solx-5,solx+5,1000)
    line1=a0+mtransform(xs,a1,1)
    line2=b0+mtransform(xs,b1,1)
    line1=frame(line1,index=xs,columns=[eq1])
    line2=frame(line2,index=xs,columns=[eq2])
    if sol:
        sy=round(B[0],3)
        sx=round(B[1],3)
        line1.plot(title= eq1 + ' and  ' + eq2 + '\n Solution:'+ 'y=' +str(sy)+ '  x=' + str(sx),grid=True,linewidth=2,ax=ax)
    else:
        sy=B[0]
        sx=B[1]
        line1.plot(title= eq1 + ' and  ' + eq2 + ':'+ sy + str(sx),grid=True,linewidth=2,ax=ax)
    line2.plot(grid=True,linewidth=2,ax=ax)
    ax.set_xlabel(lx,fontsize=14)
    ax.set_ylabel(ly,fontsize=14)
    ax.plot(solx,sy,'ob')
    ax.vlines(sx,0,sy, linestyles="dashed")
    ax.hlines(sy,0,sx, linestyles="dashed")
    plt.show()
    return

def plotfunction(c0=0,c1=1,p1=1,c2=0,p2=0,c3=0,p3=0,span=(-1,1),lx='x',ly='y',der=True):
    eq=textder(c0,c1,p1,c2,p2,c3,p3, y=ly, x=lx)
    deq=textder(c0,c1,p1,c2,p2,c3,p3,y=ly, x=lx,der=der)
    xs=np.linspace(span[0],span[1],1000)
    
    y=c0+mtransform(xs,c1,p1)+mtransform(xs,c2,p2)+mtransform(xs,c3,p3)
    c0,c1,p1,c2,p2,c3,p3=deriv(c0,c1,p1,c2,p2,c3,p3)
    z=c0+mtransform(xs,c1,p1)+mtransform(xs,c2,p2)+mtransform(xs,c3,p3)  
    
    if der==True:
        fig,ax=plt.subplots(nrows=2)
        y=frame(y,index=xs,columns=[eq])
        z=frame(z,index=xs,columns=[deq])
        x=frame(zeros(1000),index=xs)
        y.plot(title= eq,grid=True,ax=ax[0],linewidth=3)
        z.plot(title=deq,grid=True,ax=ax[1],linewidth=3, color='red')
        x.plot(grid=True,ax=ax[1],linewidth=3, color='black')
        ax[0].set_ylabel(ly,fontsize=16)
        ax[1].set_xlabel(lx,fontsize=16)
    else:
        y=frame(y,index=xs,columns=[eq])
        x=frame(zeros(1000),index=xs)
        y.plot(title= eq,grid=True,linewidth=3)
    plt.show()    
    return

def tderiv(c,p):
   
    if p =='ln':
        p=-1
    elif p=='exp':
        p='exp'  
    else:
        c=c*p
        if p != 0:
            p=p-1
    return c,p
    
def deriv(c0,c1,p1,c2,p2,c3,p3):
        c0=0
        c1,p1=tderiv(c1,p1)
        c2,p2=tderiv(c2,p2)
        c3,p3=tderiv(c3,p3)
        c0,c1,p1,c2,p2,c3,p3=amalg(c0,c1,p1,c2,p2,c3,p3)
        c0=round(c0,6)
        c1=round(c1,6)
        c2=round(c2,6)
        c3=round(c3,6)
        return  c0,c1,p1,c2,p2,c3,p3

def minimise(f,x=0):
    h=minimize(f,x)
    #print(h.x,h.fun)
    return h

def maximise(f,x=0):
    def g(x): return -f(x)
    h=minimize(g,x)
    h.fun=-h.fun
    #print(h.x,h.fun)
    return h

def plotMP(c0=0,c1=1,c2=0,c3=0,r1=0,r2=0,span=np.nan,lx='x',der=True):

    
    def profit(x,c=(c0,c1,c2,c3,r1,r2)):
         y=c[0]+mtransform(x,c[1],1)+mtransform(x,c[2],2)+mtransform(x,c[3],3)
         v=mtransform(x,c[4],1)+mtransform(x,c[5],2)
         return v-y
       
    maxp=maximise(profit,1)
    maxp.x=round(maxp.x,4)
    maxp.fun=round(maxp.fun,4)
    
    full=True
    if r1==0 and r2==0:
        maxp.x=0
        maxp.fun=0    
        full=False
        
    def avcost(x,c=(c0,c1,c2,c3)):
        y=c0+mtransform(x,c1,1)+mtransform(x,c2,2)+mtransform(x,c3,3)
        ac=y/x
        return ac
    
    minac=minimise(avcost,1)
    minac.x=round(minac.x,4)
    minac.fun=round(minac.fun,4)
    
    if np.isnan(span).any():
        if maxp.x>0:
            span=(maxp.x/10,maxp.x*2)
        else:
            span=(minac.x/10,minac.x*2)
            
            
    eqc= textder(c0,c1,1,c2,2,c3,3, y='C', x=lx)
    
    deqc=textder(c0,c1,1,c2,2,c3,3,y='C',  x=lx,der=der)
    eqr= textder(0,r1,1,r2,2, 0,3, y='R', x=lx)
    deqr=textder(0,r1,1,r2,2, 0,3, y='R',  x=lx,der=der)
    eqp= textder(r1,r2,1,0,0,0,    y='P', x=lx)
    xs=np.linspace(span[0],span[1],1000)
    
    y=c0+mtransform(xs,c1,1)+mtransform(xs,c2,2)+mtransform(xs,c3,3)
    ac=y/xs
    
    c0,c1,p1,c2,p2,c3,p3=deriv(c0,c1,1,c2,2,c3,3)
    z=c0+mtransform(xs,c1,p1)+mtransform(xs,c2,p2)+mtransform(xs,c3,p3)  
    
    v=mtransform(xs,r1,1)+mtransform(xs,r2,2)+mtransform(xs,0,3)
    r0,r1,p1,r2,p2,r3,p3=deriv(0,r1,1,r2,2,0,3)
    u=r0+mtransform(xs,r1,1)+mtransform(xs,r2,p2)+mtransform(xs,r3,p3)
    
    if der==True:
        fig,ax=plt.subplots(nrows=2)
        y=frame(y,index=xs,columns=[eqc])
        z=frame(z,index=xs,columns=[deqc])
        v=frame(v,index=xs,columns=[eqr])
        u=frame(u,index=xs,columns=[deqr])
        ac=frame(ac,index=xs,columns=['AC'])
        
        x=frame(zeros(1000),index=xs)
        
        y.plot(title= '',grid=True,ax=ax[0],linewidth=3)
        if not full:
            v.plot(title= 'Cost and Revenue' + '\n Acost MinX=' +str(minac.x) + ' AvCost=' +str(minac.fun),grid=True,ax=ax[0],linewidth=3)
        else:
            v.plot(title= 'Cost and Revenue: where ' + eqp + '\nProfit Max X=' +str(maxp.x) + ' Profit=' +str(maxp.fun) + '\n Acost MinX=' +str(minac.x) + ' AvCost=' +str(minac.fun),grid=True,ax=ax[0],linewidth=3)
        
        z.plot(title= '',grid=True,ax=ax[1],linewidth=3, color='red')
        ac.plot(title= '',grid=True,ax=ax[1],linewidth=3, color='black')
        u.plot(title= 'Marginal Revenue and Cost',grid=True,ax=ax[1],linewidth=3, color='blue')
        
        #x.plot(grid=True,ax=ax[1],linewidth=3, color='black')
        ax[0].set_ylabel('',fontsize=16)
        ax[1].set_xlabel(lx,fontsize=16)
    else:
        y=frame(y,index=xs,columns=[eqc])
        x=frame(zeros(1000),index=xs)
        y.plot(title= eqc,grid=True,linewidth=3)
        v=frame(y,index=xs,columns=[eqr])
        v.plot(title= eqr,grid=True,linewidth=3)
    plt.show()    
       
    return maxp,minac
    
def deriv(c0,c1,p1,c2,p2,c3,p3):
        c0=0
        c1,p1=tderiv(c1,p1)
        c2,p2=tderiv(c2,p2)
        c3,p3=tderiv(c3,p3)
        c0,c1,p1,c2,p2,c3,p3=amalg(c0,c1,p1,c2,p2,c3,p3)
        c0=round(c0,6)
        c1=round(c1,6)
        c2=round(c2,6)
        c3=round(c3,6)
        return  c0,c1,p1,c2,p2,c3,p3

def amalg(c0,c1,p1,c2,p2,c3,p3):        
        if p1==0:
           c0=c0+c1
           c1=0
        if p2==0:
           c0=c0+c2
           c2=0 
        if p3==0:
           c0=c0+c3
           c3=0   
        return c0,c1,p1,c2,p2,c3,p3
    
def textder(c0=0,c1=1,p1=0,c2=0,p2=0,c3=0,p3=0, x='x',y='y',der=False):
    c0,c1,p1,c2,p2,c3,p3=amalg(c0,c1,p1,c2,p2,c3,p3)
    
    if der:
       c0,c1,p1,c2,p2,c3,p3=deriv(c0,c1,p1,c2,p2,c3,p3) 
       #print('yes', c0,c1,p1,c2,p2,c3,p3)
       #stop()
    beg=r''+y+'= '
    dbeg=r'd'+y+'/d'+x+'= '
    #dbeg=r'dy/dx= '
    if der==False:   
        s=beg
    else:
        s=dbeg
        
    if c0==0:
       add=''
    else:
       add='+' 
    
    if c0 !=0:
       s=s+ str(c0) 
       
    s=augment(s,x,c1,p1,add)
    s=augment(s,x,c2,p2,'+') 
    s=augment(s,x,c3,p3,'+') 
    
    if s==beg:
       s==beg+ str(0)
    
    if s==dbeg:
       s= dbeg + str(0)
    
    return s    

def augment(s,x,c,p,add):    
    if c !=0:
        if c>0:
            if p==0:
                s=s + add + str(c)
            else:    
                s=s + add + str_i(c,p)+ powerov(x,p) 
        if c<0: 
            if p==0:
                s=s + ' ' + str(c) 
            else:
                s=s + ' ' + str_i(c,p)+ powerov(x,p) 
    return s  

def str_i(c,p):
    
    if p!=0:
       if (c == 1 or c==1.0 ) :
           ret=' '
       elif (c==-1 or  c==-1.0):  
           ret= '-'
       else:
           ret=c
    else:
       ret=c 
    return str(ret) 

def solveline(x0,y0,x1,y1):
    slope=(y1-y0)/(x1-x0)
    intercept=y0-slope*x0
    return intercept,slope

def powerov(x='x',p=1):
    if p==1:
        s= x
    elif p==0:
        s='1'    
    elif p=='ln':
        s='ln(x)'
    elif p=='exp':
        s='exp(x)'    
    else:    
        s= '$'+ x +'^' + str({p}) + '$'
    return s

def mtransform(x,c,p):
    if p=='ln':
       z=c*ln(x)
    elif p=='exp':
       z=c*exp(x)
    else:
       z=c*x**p
    return z   


def printlong(x):
    pd.set_option('display.max_rows', rows(x))
    print(x)
    pd.reset_option('display.max_rows')   

def printwide(x):
    pd.set_option('display.max_columns', cols(x))
    print(x)
    pd.reset_option('display.max_columns')   
    
    
def rankdatap(x,method='dense',ascend=True):
    x=twodm(x)
    y=twodm(sps.rankdata(x,method))
    if ascend:
        y=np.max(y)-y+1
    return cc([x,y])

def zscore(x):
    return (x-meanc(x).T)/stdc(x).T

def idiag(x):
    n=np.max(shape(x))
    y=np.eye(n)*x
    return y

def rankdatap(x,method='dense',ascend=True):
    x=twodm(x)
    y=twodm(sps.rankdata(x,method))
    if ascend:
        y=np.max(y)-y+1
    return cc([x,y])

def eigh(c):
    val,vc=np.linalg.eigh(c)
    #print(val,vc)
    idx = (rankdatap(val)[:,1]).astype(int)-1
    val=twodma(val)
    vc=twodma(vc)
    val = val[idx]
    vc = vc[:,idx]
    return val,vc

def princomp_(x,typ='corr'):
    x=x-meanc(x).T
    if typ=='cov':
        val,vc=eigh(cov(x))
    else:
        x=zscore(x)        
        val,vc=eigh(cov(x))
    svc=vc.dot(sqrt(idiag(val**-1)))
    scores=x.dot(vc) #scores
    nscores=x.dot(svc) #normalised scores
    
    val=findex(frame(val,columns=['eigv']))
    val['cproportion']=val.cumsum()
    val['cproportion']=val['cproportion']/float(sumc(val['eigv']))
    val['proportion']=val['eigv']
    val['proportion']=val['eigv']/float(sumc(val['eigv']))
    
    return [val,findex(findex(vc).T).T,findex(findex(svc).T).T,findex(findex(scores).T).T,findex(findex(nscores).T).T]       

def princomp(data,y='',typ='corr',n=0,varimax=True):
    '''
    a[0] #Eigenvalues etc
    a[1] #Raw Factor Loadings 
    a[2] #Normalised Factor Loadings
    a[3] #Unormalised PC
    a[4] #Normalised  PCs
    a[5] #Communalities, if n>0
    a[6] #Normalised Varimax Rotated Factors
    a[7] #Rotation
    a[8] #Rotated Factor Loadings
    a[9] #Communalities Rotated if n>0
    '''
    dats=data[y].copy().dropna()
    dat=twodma(dats)
    dat=dat-meanc(dat).T
    if typ=='cov':
        dat=frame(dat)
        dat.index=dats.index
        dat.columns=dats.columns
    else:
        dat=frame(zscore(dat))  
        dat.index=dats.index
        dat.columns=dats.columns
    a=princomp_(dat,typ=typ)
    a[1].index=y
    a[2].index=y
    a[3].index=dats.index
    a[4].index=dats.index
    if n !=0:
        a=a+[coms(dat,y=y,pcs=a[3],n=n)]
    else:
        a=a+[frame()]
    a[3]=cc([data.iloc[:,0],a[3]])
    del a[3][a[3].columns[0]]
    a[4]=cc([data.iloc[:,0],a[4]])
    del a[4][a[4].columns[0]]
    if varimax:  
       Phi,R=Varimax(a[4]) 
       a=a+[findex(Phi.T).T]  #Rotated Normalised Factors, in a[5]
       a=a+[frame(R)]         #The Rotation matrix in a[6]
       a=a+[findex(frame(twodma(a[2]).dot(R),index=a[2].index).T).T] #Rotating the factor loadings
       #print(frame(twodma(a[1].T).dot(R.T)))
       #print(frame(twodma(a[1]).dot(R)))
    else:
       a=a+[frame()]
       a=a+[frame()]
       a=a+[frame()]
    if n !=0:
        a=a+[coms(dat,y=y,pcs=Phi,n=n)]
    else:
        a=a+[frame()]    
    return a

def Varimax(Phi, gamma = 1, q = 500, tol = 1e-6):
    #Phi should be res[2] or res[3] from princomp
    #from numpy import eye, asarray, dot, sum, diag
    from numpy.linalg import svd
    p,k = Phi.shape
    R = np.eye(k)
    d=0
    for i in range(q):
        d_old = d
        Lambda = np.dot(Phi, R)
        #Lambda = Phi.dot(R)
        u,s,vh = svd(np.dot(Phi.T,np.asarray(Lambda)**3 - (gamma/p) * np.dot(Lambda, np.diag(np.diag(np.dot(Lambda.T,Lambda))))))
        R = np.dot(u,vh)
        #R = u.dot(vh)
        d = sum(s)
        if d/d_old < tol: 
            print('Yes',i) 
            break
    #return np.dot(Phi, R)
    #print(i,d/d_old)
    return np.dot(Phi,R),R


def coms(data,y='y',pcs='pcs',n='n'):
    x=cc([ones(rows(data)),twodma(pcs)[:,0:n]])
    P=x.dot(inv(x.T.dot(x))).dot(x.T)
    M=np.eye(rows(x))-P
    rec=list()
    for i in y:
        ys=zscore(twodma(data[i]))
        y_=twodma(ys-meanc(ys))
        rss=float(ys.T.dot(M).dot(ys))
        ess=float(ys.T.dot(P).dot(ys))
        tss=float(y_.T.dot(y_))
        rec=rec+[float(ess/tss) ]
    return frame(rec,index=y,columns=['Communalities'])


def myOLS(y,x,printit=True):    
    res=OLS(y,x).fit()
    if printit=='simple':
            res.out=sumry(res)
    if printit==True:    
            print(res.summary2())
            '''
            test1 = breushpagan(res.resid, res.model.exog)
            print('\nbresuchpagan',test1)     
            test2=breushgodfrey(res, nlags=1, store=False)
            print('\nbresuchgodfrey',test2)
            '''
    return res

def regression(data):
    print('Your variables are:', list(data.columns))
    print('Note that regression variables cannot have gaps (white spaces) in their names')
    return interact_manual(regols,data=fixed(data),formula='',robust=False,printit=['simple','full'],BPtest=False,BGtest=False,LOS=[0.05,0.10,0.025,0.01])

def regols(data,formula,robust=False,printit='simple',rank=False,ploterror=False,anova_table=False,BPtest=False,BGtest=False,LOS=0.05,hypotheses=''):
    LOS=float(LOS)
    formula=formula.replace('=','~')
    try:
        datas=data.copy().astype(float)
    except:
        datas=data.copy()
    try:
        datas[formula.split('~')[0]]=data[formula.split('~')[0]].astype(float)  
    except:
        pass
    if rank==True:
        datas[formula.split('~')[0]]=rankdata(data[formula.split('~')[0]])
    if robust:
        robust_='HC3'
    else:
        robust_='nonrobust'
    res = ols(formula,datas).fit(cov_type=robust_)
    if printit=='simple':
            res.out=sumry(res,LOS)
    else:
        print(res.summary())
    if ploterror:    
        plothist(frame(res.resid),normal=True) 
        plt.show()
        plotbox(frame(res.resid))
        plt.show()
    if anova_table:
        print(anova(res,typ=2))
    if BPtest:    
        test1 = breushpagan(res.resid, res.model.exog)
        test1=array(test1)
        test1=frame(reshape(test1,2,2),columns=['F','LM'],index=['val','pval'])
        print('\nBresuch-Pagan test for Heteroscedasticity\n',test1.round(4))   
    if BGtest:    
        test2=breushgodfrey(res, nlags=1, store=False)
        test2=array(test2)
        test2=frame(reshape(test2,2,2),columns=['F','LM'],index=['val','pval'])
        print('\nBresuch-Godfrey for First order Serial Correlation\n',test2.round(4))   
    if hypotheses !='':
        out= res.f_test(hypotheses)
        
        f_test=frame([float(out.fvalue),float(out.pvalue),float(out.pvalue)<LOS],index=['F-value','P-Value','Reject at ('+str(100*LOS)+ '% LOS)'],columns=['F-test'])
        'Reject 0 ('+str(100*LOS)+ '% LOS)'
        print('\nResults of the hypothesis test for: '+ hypotheses+':\n',f_test.T)
    return res


def regols_(data,formula,robust='nonrobust',printit='simple'):
    try:
        datas=data.copy().astype(float)
    except:
        datas=data.copy()
    try:
        datas[formula.split('~')[0]]=makenormal(data[formula.split('~')[0]])
    except:
        pass
    if robust==False:
        robust='nonrobust'
    #datas=datas.dropna()
    res = ols(formula,datas).fit(cov_type=robust)
    if printit=='simple':
            res.out=sumry(res)
    if printit==True:
        print(res.summary())
    return res


def sumry(resm,LOS):
    print('____________________________________________________________________')
    print("Rsquared", resm.rsquared.round(6))
    rp= (resm.rsquared*100).round(2)
    print("(", rp,"% of the variation in the dependent variable can be explained by the independent variables)")
    print('____________________________________________________________________')
    rf=round(resm.f_pvalue*100+.5)
    fvalue=frame([squeeze(resm.fvalue),squeeze(resm.f_pvalue)],index=['F-Value','P-Values'],columns=['No Regression'])
    print('\nTest for Null (H0): ALL the regressors having a zero slope (i.e. No Regression)')
    print(fvalue.T.round(6))
    if resm.f_pvalue< LOS:
        print('The NULL (H0) CAN be rejected at the', str(100*LOS) , '% LOS  in favour of the hypothesis that at least one \nof regressors has a non-zero coefficient')
    else:
        print('The NULL (H0) CANNOT be rejected at the' , str(100*LOS) , '% LOS')
    
    print('____________________________________________________________________')
    print("\nModel Coefficients and Significance")
    out1=frame(resm.params)
    out2=frame(resm.pvalues)
    out3=frame(resm.bse)
    out4=frame(resm.tvalues)
    out5=frame(resm.pvalues<LOS)
    
    out=cc([out1,out2,out3,out4,out5])
    out.columns=['Coef','P-Vals','Std.errs','t','Reject 0 ('+str(100*LOS)+ '% LOS)']        
    print(out.round(6))
    print('Note: that the tests and p-values are for two tailed alternatives') 
    return

def split(text,delimiter):
    x=text.split(delimiter)
    j=0
    y=[]
    for i in x:
        if j< len(x)-1:
            y= y+ [i] + [delimiter]
        else:
            y= y+ [i]
        j=j+1
    return y  

def colorit(data,xname,xvalue,color1='black',color2='red'):
    colors=list()
    for i in range(rows(data)):
        if data.loc[i,xname]==xvalue:
            #print(i)
            colors=colors+['black']
        else:    
            #print(i)
            colors=colors+['red']
    return colors

def regdraw(data,formula,regressor=1,colors='blue',dummy='',rank=False):
    datas=data.copy()
    if rank==True:
        datas[formula.split('~')[0]]=rankdata(data[formula.split('~')[0]])
    y,x=patsy.dmatrices(formula,datas)
    namesx=x.design_info.column_names
    namesy=y.design_info.column_names
    namex=namesx[regressor]
    namey=namesy[0]
      
    res1=regols(datas,formula,rank=rank)
    
    b=res1.params
    pvals=frame(res1.pvalues)
    pvals.columns=['P']
    
    
    labs1=split(formula,'~')
    labs2=split(labs1[2],'+')
    labs3=labs1[0] + ' = ' + str(round(b[0],3)) 
    
    j=0
    for i in range(1,rows(b)):  
        #q=round(twodma([b[i],pval[i]]),4)
        q=round(b[i],3)
        if b[i]>0:
            labs3=labs3 + ' + '  + str(q) + ' * ' + labs2[j]
        else:
            labs3=labs3 + ' - '  + str(abs(q)) + ' * ' + labs2[j]
        j=j+2
    labs3=labs3+ '\n\n Rsquare =' + str(round(res1.rsquared,3))  
    labs3=labs3+ '\n\n' + str(pvals.T.round(4))
    

    fig, ax = plt.subplots()
    #fig.suptitle(labs3,bbox=props )    
    
    if dummy=='': 
        y0,x0=predline(x,b,regressor)
        ax.plot(squeeze(x0), squeeze(y0),color='black')
    else:
        y0,x0=predline(x,b,regressor,t=[dummy,0])
        y1,x1=predline(x,b,regressor,t=[dummy,1])
        ax.plot(squeeze(x0), squeeze(y0), color='black')
        ax.plot(squeeze(x0), squeeze(y1), color='black')
    ax.set_xlabel(namex)
    ax.set_ylabel(namey)
    props = dict(boxstyle='square', facecolor='r', alpha=0.1)
    #ax.text(minc(x0),maxc(y),str(pvals.round(4)),,fontsize=12,bbox=props)
    ax.text(minc(x0),maxc(y),labs3,verticalalignment='top',fontsize=12,bbox=props)
    ax.scatter(squeeze(x0), y, color=colors)
    plt.grid()
    fig.show()
    return res1


def predline(x,b,s,t=''):
    if t !='':
       x[:,t[0]]=t[1]
    s=[s]
    t=list()
    for i in range(rows(b)):
        t=t +[i]
    u=list(set(t)-set(s)) #Those set constant
    z=twodma(x[:,u])      
    z=meanc(z).T
    x=twodma(x)
    v=x[:,s].dot(twodma(b[s])) #those that varying
    q=z.dot(twodma(b[u]))      #those held constant
    y1=v+q
    return y1,x[:,s]

def gforc(y,b):
    ynew=b[0]
    for i in range(1,len(b)):
        ynew=ynew+b[i]*y[i]
    return ynew

def fforc(y,b):
    y0=gforc(y,b)
    z=[y0]
    for i in range(1,len(y)):
        z=z+[y[i-1]]
    return z    

def forecast(dats,periods,res):
    b=res.params
    n=rows(dats)-1
    y=dats.iloc[n]
    forc=list()
    for i in range(periods):
        y=fforc(y,b)
        forc=forc+[y[0]]
    forc=frame(forc)
    forc.columns=['forecast']
    return forc

def regprobit(data,formula,printit=True):
    formula=formula.replace('=','~')
    res = smf.probit(formula,data).fit()    
    if printit==True:
        print(res.summary2())
        mfx = res.get_margeff()
        print(mfx.summary()) 
    return res 

def probit(data):
    print('Your variables are:', list(data.columns))
    print('Note that regression variables cannot have gaps (white spaces) in their names')
    return interact_manual(regprobit,data=fixed(data),formula='',printit=fixed(True))

def ordprobit(data):
    print('Your variables are:', list(data.columns))
    print('Note that regression variables cannot have gaps (white spaces) in their names')
    return interact_manual(regoprobit,data=fixed(data),formula='',printit=fixed(True))


def findcats(y):
    y=twodma(y)
    h=list()
    for i in range(10):
        s=sumc(twodma(y)==i)
        if s>0:
           h=h+[i]
    return h    

def oprobit_lik(y, X, beta, lamda,h):
    mu = twodma(np.dot(X, beta))
    j=1
    ll=0
    for i in h:
        s=y==i
        mui=mu[s]
        if j==1:
           l=norm.cdf(0,mui,1.0)
        elif j==2:
           l=norm.cdf(lamda[0],mui,1.0)- norm.cdf(0,mui,1.0)
        elif j<len(h):
           l=norm.cdf(lamda[j-2],mui,1.0)- norm.cdf(lamda[j-3],mui,1.0)
        else:
           l=norm.sf(lamda,mui,1)
        ll=ll+ln(l).sum()
        j=j+1
    return ll
    
    
class oprobit(GenericLikelihoodModel):
    def __init__(self, endog, exog, **kwds):
        self.h=kwds['h']
        super(oprobit, self).__init__(endog, exog, **kwds)

    def nloglikeobs(self, params):
        k=len(self.h)-2
        lamda = params[-k:]
        beta = params[:-k]
        ll = oprobit_lik(self.endog, self.exog, beta, lamda,self.h)
        return -ll

    def fit(self, start_params=None, maxiter=100000, maxfun=50000, **kwds):
        # we have one additional parameter and we need to add it for summary
        #self.exog_names.append('lambda')
        if start_params == None:
            # Reasonable starting values
            lx=np.array(self.exog)
            ly=np.array(self.endog)
            start_params0=(inv(lx.T.dot(lx))).dot(lx.T.dot(ly))
            
            bnd=0.5
            start_params = np.append(start_params0, bnd)
            u=len(self.h)-3
            bnd=bnd+0.5
            for i in range(u):
                start_params = np.append(start_params, bnd)
                bnd=bnd+0.5
            #start_params = np.zeros(4)
            #print('yes')
            print(start_params)
            # intercept
            #start_params[-2] = np.log(self.endog.mean())
        return super(oprobit, self).fit(start_params=start_params,
                                     maxiter=maxiter, maxfun=maxfun,
                                     **kwds)

def myOPROBIT(y,x,printit=True):
    h=findcats(y)
    y=y.astype(float)
    res = oprobit(y,x,h=h).fit()
    if printit:
        print(res.summary())
    return res 

def regoprobit(data,formula,printit=True):
    formula=formula.replace('=','~')
    y_,x=patsy.dmatrices(formula,data)
    #print(rows(y_),cols(y_))
    y=list()
    '''
    for i in range(rows(y_)):
        for j in range(cols(y_)):
            if y_[i,j]==1.0:
               y=y+[j]  
    '''
    y=frame(y_)       
    x=frame(x,columns=[x.design_info.column_names])
    #print(y)
    #print(x)
    res=myOPROBIT(y,x,printit)
    return res


def search(k,s=dir()):
    v=list()
    for i in s:
        if k in i:
            v=v+[i]    
    return v

def statew(y='y',x='',mu='',groupby='',group1='',group2=''):
    s0=x != ''
    s1=mu != ''
    s2=groupby != ''
    t2=groupby==''
    s3=group1 != ''
    t3=group1==''
    s4=group2 != ''
    t4=group2==''
    return s0,s1,s2,s3,s4,t2,t3,t4

    
def warn(y='y',x='',mu='',groupby='',group1='',group2=''):
    s0,s1,s2,s3,s4,t2,t3,t4=statew(y,x,mu,groupby,group1,group2)

    if s0:
        if s1|s2|s3|s4:
            print('Since x is active, the other settings (mu, groupby,group1, group2) are treated as blank')
            mu='';groupby='';group1='';group2=''
            s0,s1,s2,s3,s4,t2,t3,t4=statew(y,x,mu,groupby,group1,group2)
    if s1:
        if s2|s3|s4:
            print('Since mu is active and x is blank, the groupby settings are ignored')
            groupby='';group1='';group2=''
            s0,s1,s2,s3,s4,t2,t3,t4=statew(y,x,mu,groupby,group1,group2)
    if s2:
       if t3|t4:
          print('Since groupby is active, you need to specify the two groups in group1 and group2')
    if s3|s4:
       if t2:
          print('A group is specified but no groupby variable. Specify one if you wish to do a test by groups')  
    return y,x,mu,groupby,group1,group2


def query(data,args,simple=False):
    
    if type(args) !=list:
       args=args.split(' ')
    s1=query_(data,args[0],args[1],args[2],True)
    s=s1
    if len(args)>3:
        s2=query_(data,args[4],args[5],args[6],True)
        if args[3] =='or':
           s=s1|s2
        else:
           s=s1&s2
    if len(args)>7:
        s3=query_(data,args[8],args[9],args[10],True)
        if args[7] =='or':
           s=s|s3
        else:
           s=s&s3
    if len(args)>11:
        s4=query_(data,args[12],args[13],args[14],True)
        if args[11] =='or':
           s=s|s4
        else:
           s=s&s4
    if len(args)>15:
        s5=query_(data,args[16],args[17],args[18],True)
        if args[11] =='or':
           s=s|s5
        else:
           s=s&s5
        
    if simple==True:
       return np.array(s)
    else:
       return data[s]
    
def floats(x):
    try: 
        return float(x)
    except:
        print('The query contained a mathematical comparsion which required a numberic quantity for comparison')

        
def querynan(data,column,mode='is',criteria='nan'):
    datas=data[column].copy().fillna(-np.pi)
    if mode=='is':
        if criteria=='nan':
            s=datas==-np.pi
        else:
            s=datas!=-np.pi
    if mode=='isnt' or mode=='isnot':
        if criteria=='nan':
            s=s=datas!=-np.pi
        else:
            s=datas==-np.pi
    
    if mode=='eq' or mode=='==':
        if criteria=='nan':
            s=s=datas==-np.pi
        else:
            s=datas!=-np.pi
    
    if mode=='ne' or mode=='!=':
        if criteria=='nan':
            s=s=datas!=-np.pi
        else:
            s=datas==-np.pi
    return s

def query(data,args,simple=False,lower=False):
        if type(args) !=list:
           args=args.split(' ')
        s1=query_(data,args[0],args[1],args[2],True,lower=lower)
        s=s1
        if len(args)>3:
            s2=query_(data,args[4],args[5],args[6],True)
            #print(s1,s2)
            if args[3] =='or':
               s=s1|s2
            else:
               s=s1&s2
        if len(args)>7:
            s3=query_(data,args[8],args[9],args[10],True)
            if args[7] =='or':
               s=s|s3
            else:
               s=s&s3
        if len(args)>11:
            s4=query_(data,args[12],args[13],args[14],True)
            if args[11] =='or':
               s=s|s4
            else:
               s=s&s4
        if len(args)>15:
            s5=query_(args[16],args[17],args[18],True)
            if args[11] =='or':
               s=s|s5
            else:
               s=s&s5
        
        if simple==True:
           return s
        else:
           return data[s]
    
def query_(data,column,mode,criteria,simple=True,lower=False):
    #print(criteria,type(criteria))
    if mode=='in':
        mode='includes'
    x=data[column]
    if lower:
        x=x.str.lower()
    s=[]
    #The following is for strings , compare is, includes, starts,
    if mode=='is':
        if criteria=='nan':
            s=querynan(data,column)
        elif criteria=='notnan':
            s=querynan(data,column,criteria='notnan')
        else:
            s=x==criteria
    elif mode=='isnt' or mode=='isnot':
        if criteria=='nan':
            s=querynan(data,column,mode='isnot',criteria='nan')
        elif criteria=='notnan':
            s=querynan(data,column,mode='isnot',criteria='notnan')
        else:
            s=x!=criteria
    elif mode=='includes':
        for i in x:
            if criteria in i:
                s=s+[True]
            else:
                s=s+[False]
    elif mode=='excludes':
        for i in x:
            if criteria not in i:
                s=s+[True]
            else:
                s=s+[False]
    elif mode=='starts' or mode=='begins':
        n=len(criteria)
        for i in x:
            if  criteria == i[:n]:
                s=s+[True]
            else:
                s=s+[False]
    elif mode=='ends':
        n=len(criteria)
        for i in x:
            if  criteria == i[-n:]:
                s=s+[True]
            else:
                s=s+[False]
    elif mode=='eq' or mode =='==':
            if criteria=='nan':
                s=querynan(data,column,mode='eq',criteria='nan')
            elif criteria=='notnan':
                s=querynan(data,column,mode='eq',criteria='notnan')
            else:
                s=x==floats(criteria)
    elif mode=='ne' or mode =='!=':
            if criteria=='nan':
                s=querynan(data,column,mode='ne',criteria='nan')
            elif criteria=='notnan':
                s=querynan(data,column,mode='ne',criteria='notnan')
            else:
                s=x!=floats(criteria)
    elif mode=='gt' or mode=='>':
            s=x>floats(criteria)
    elif mode=='gte' or mode=='>=' or mode=='ge':
            s=x>=floats(criteria)
    elif mode=='lt' or mode=='<':
            s=x<floats(criteria)
    elif mode=='lte'or mode=='<=' or mode=='le':
            s=x<=floats(criteria)
    elif mode=='at':
            s=x==criteria
    elif mode=='notat':
            s=x!=criteria
    elif mode=='before':
            s=x<criteria  
    elif mode=='after':
            s=x>criteria          
    if simple:
        return np.array(s)
    else:
        return data[s]
    
def date(year,month=1,day=1):
    return datetime(year,month,day)

def taxcalc(amount=0,thresholds=[10],rates=[0,.1]):
    if len(rates) !=len(thresholds)+1:
       print('There need to be one more rate than threshold')
    
    
    x=array(thresholds)>=(amount)
    x=list(x)+[True]
    tax=[]
    rat=[]
    for i in range(len(thresholds)):
        a=x[i]
        b=x[i+1]
        if ~a:
           rat=rat+[rates[i]]
           tax=tax+[thresholds[i]]
        if ~a and b:
           rat=rat+[rates[i+1]] 
           tax=tax+[amount]

    if len(tax)==0:
        d=frame([amount,rates[0]]).T
        d.columns=['Increments','Tax Rate']
    else:
        inc=[tax[0]]
        r=[rates[0]]
        for i in range(1,len(tax)):
            #print(i)
            inc=inc+[tax[i]-tax[i-1]]
            r=r+[rates[i]]
            d=frame(inc,columns=['Increments'])
            d['Tax Rate']=r
        
    d['Gross Income']=d['Increments'].cumsum()
    d['Tax']=d['Increments']*d['Tax Rate']
    d['Total Tax']=d['Tax'].cumsum()
    d['Net Income']=d['Gross Income']-d['Total Tax']
    d=d[['Gross Income','Increments','Tax Rate','Tax','Total Tax','Net Income']]
    return d

def mergerates(xs,zs,rs,ts):
    x=xs.copy()
    z=zs.copy()
    r=rs.copy()
    t=ts.copy()
    X=[]
    R=[]
    mz=0
    mx=0
    while len(x)>0 or len(z)>0:
       if len(x)>0 and len(z)>0: 
           if x[0]<z[0]:
               h=x.pop(0)
               mx=r.pop(0);mz=t[0]
               X=X+[h]
               m=mz+mx  
               R=R+[m] 
           else:
               h=z.pop(0)
               mz=t.pop(0);mx=r[0]
               X=X+[h]
               m=mz+mx 
               R=R+[m] 
       elif len(x)>0:
            h=x.pop(0)
            mx=r.pop(0);mz=t[0]
            X=X+[h]
            m=mz+mx 
            R=R+[m] 
            state=True
       elif len(z)>0:
            h=z.pop(0)
            z=t.pop(0);mx=r[0]
            X=X+[h]
            m=mz+mx 
            R=R+[m]
            state=False
            #print(x,z) 
    rr=r[0]+t[0]
    if rr>1:
       rr=1
    R=R+[rr]    
    z=t.pop(0);mx=r[0]
    return X,R
#The following are for Input Output Analysis

def stepa(IM,FD):
    IM=twodma(IM)
    FD=twodma(FD)
    ID=sumc(IM.T)
    TD=FD+ID
    A=IM/TD.T
    I=np.eye(rows(A))
    M=inv(I-A);
    return A,M,ID,FD,TD

def frames(IM,HD,ED,sectors):
    FD=array(HD)+array(ED)
    A,M,ID,FD,TD=stepa(IM,FD)
    IM_=frame(IM,index=sectors,columns=sectors)
    HD_=frame(HD,index=sectors,columns=['Household Demand'])
    ED_=frame(ED,index=sectors,columns=['Exports'])
    ID_=frame(ID,index=sectors,columns=['Intermediate Demand'])
    FD_=frame(FD,index=sectors,columns=['Final Demand'])
    TD_=frame(TD,index=sectors,columns=['Total Demand'])
    A_=frame(A,index=sectors,columns=sectors)
    M_=frame(M,index=sectors,columns=sectors)
    Demands=cc([ID_,HD_,ED_,FD_,TD_])
    return IM_,Demands,A_,M_

def dshocks(IM,Demands,DFD,sectors):
        FD=Demands['Final Demand']
        DFD_=frame(DFD,index=sectors,columns=['Demand Shock'])
        A,M,ID,FD,TD=stepa(IM,FD)
        I=np.eye(rows(A))
        NFD=FD.copy()
        NFD[0]=NFD[0]+DFD[0]
        NFD[1]=NFD[1]+DFD[1]
        NFD[2]=NFD[2]+DFD[2]
        NFD_=frame(NFD,columns=['New Final Demand'],index=sectors)
   
        NTD=M.dot(NFD);  NTD_=frame(NTD,columns=['New Total Demand'],index=sectors)
        NIM=A.dot(I*NTD);NIM_=frame(NIM,columns=sectors,index=sectors)
        TD_=frame(TD,index=sectors,columns=['Total Demand Original'])
        
        Demand=cc([DFD_,TD_,NTD_])
        Demand['Change']=Demand['New Total Demand']-Demand['Total Demand Original']
        return Demand,NIM_;
    
def resourcechange(Name,resource,Demands):
    resource=twodma(resource)
    f1=twodma(Demands['Change'])
    f2=twodma(Demands['Total Demand Original'])
    Demands[Name+' change']=f1*resource/f2
    Intensity=resource/f2
    Demands[Name+' Intensity']=Intensity
    return Demands
    
